"""Локальная сборка отчёта table_generator: Excel, Smart Chart (без OpenRouter)."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from io import BytesIO

try:
    import matplotlib

    matplotlib.use("Agg")
except ImportError:
    matplotlib = None  # type: ignore[assignment,misc]

from services.table_chart_types import ChartType
from services.table_json import TableJsonPayload, parse_table_json_response, table_payload_has_data
from services.table_markdown import (
    normalize_table_rows,
)
from services.table_xlsx_preprocess import pick_telegram_preview_rows, preprocess_xlsx_rows
from services.table_xlsx_flow import build_wb_telegram_preview_html
from services.table_text_response import (
    build_table_one_screen_html,
    build_wb_finance_express_html,
    compute_wb_marketplace_metrics,
)
from services.table_wb_chart import extract_wb_sales_series, try_render_wb_chart_png
from services.telegram_safe_text import _escape_telegram_html, repair_telegram_html

logger = logging.getLogger(__name__)

TABLE_XLSX_FILENAME = "Отчет_Нейросеть.xlsx"
_CAPTION_MAX = 1020

_TIME_AXIS_KEYWORDS: frozenset[str] = frozenset(
    {
        "год",
        "год.",
        "месяц",
        "квартал",
        "дата",
        "время",
        "период",
        "quarter",
        "year",
        "month",
        "date",
    }
)


@dataclass(frozen=True)
class TableGeneratorPack:
    rows: list[list[str]]
    telegram_caption_html: str
    xlsx_bytes: bytes
    chart_png_bytes: bytes | None
    chart_type: ChartType
    calculated_total: float = 0.0


def markdown_table_to_telegram_caption(
    rows: list[list[str]],
    *,
    title: str | None = None,
) -> str:
    rows = normalize_table_rows(rows)
    if not rows:
        heading = _escape_telegram_html(title or "Отчёт NeuroMule")
        return f"<b>📊 {heading}</b>"
    ncols = len(rows[0])
    widths = [max(len(rows[r][c]) for r in range(len(rows))) for c in range(ncols)]

    def render_row(cells: list[str]) -> str:
        padded = [cells[c].ljust(widths[c]) if c < len(cells) else "".ljust(widths[c]) for c in range(ncols)]
        return " │ ".join(padded)

    lines = [render_row(rows[0])]
    if len(rows) > 1:
        lines.append("─" * (sum(widths) + 3 * max(ncols - 1, 0)))
        for row in rows[1:]:
            lines.append(render_row(row))
    body = _escape_telegram_html("\n".join(lines))
    heading = _escape_telegram_html(title or "Отчёт NeuroMule")
    return repair_telegram_html(f"<b>📊 {heading}</b>\n<pre>{body}</pre>")


def _parse_number(raw: str) -> float | None:
    text = (raw or "").strip()
    if not text:
        return None
    text = text.replace("\u00a0", " ").replace(" ", "")
    text = re.sub(r"[₽$€%]", "", text)
    if "," in text and "." in text:
        text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)
    if not text or text in {".", "-", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _detect_chart_axes(rows: list[list[str]]) -> tuple[int, int] | None:
    rows = normalize_table_rows(rows)
    if len(rows) < 2:
        return None
    headers, data = rows[0], rows[1:]
    ncols = len(headers)
    numeric_cols: list[int] = []
    for col in range(ncols):
        nums = [_parse_number(row[col] if col < len(row) else "") for row in data]
        valid = [n for n in nums if n is not None]
        if len(valid) >= max(2, len(data) // 2):
            numeric_cols.append(col)
    if not numeric_cols:
        return None
    label_candidates = [c for c in range(ncols) if c not in numeric_cols]
    label_col = label_candidates[0] if label_candidates else 0
    value_col = numeric_cols[0] if numeric_cols[0] != label_col else numeric_cols[-1]
    if label_col == value_col:
        return None
    return label_col, value_col


def _extract_series(rows: list[list[str]]) -> tuple[list[str], list[float], str] | None:
    axes = _detect_chart_axes(rows)
    if axes is None:
        return None
    label_col, value_col = axes
    rows = normalize_table_rows(rows)
    labels: list[str] = []
    values: list[float] = []
    for row in rows[1:]:
        label = (row[label_col] if label_col < len(row) else "").strip() or "—"
        num = _parse_number(row[value_col] if value_col < len(row) else "")
        if num is None:
            continue
        labels.append(label[:40])
        values.append(num)
    if len(values) < 2:
        return None
    header = rows[0][value_col] if value_col < len(rows[0]) else "Значение"
    return labels, values, header


def _contains_time_axis_keyword(*texts: str) -> bool:
    """Ищет временные маркеры в заголовках / подписях оси (регистронезависимо)."""
    for raw in texts:
        low = (raw or "").strip().lower()
        if not low:
            continue
        for keyword in _TIME_AXIS_KEYWORDS:
            if keyword in low:
                return True
    return False


def _first_column_values(rows: list[list[str]]) -> list[str]:
    rows = normalize_table_rows(rows)
    if len(rows) < 2:
        return []
    return [(row[0] if row else "").strip() for row in rows[1:]]


def _is_strictly_sequential_integers(values: list[str]) -> bool:
    """Первая колонка — строго последовательные целые (2021, 2022, 2023 …)."""
    ints: list[int] = []
    for raw in values:
        num = _parse_number(raw)
        if num is None or abs(num - round(num)) > 1e-9:
            return False
        ints.append(int(round(num)))
    if len(ints) < 2:
        return False
    ordered = sorted(ints)
    return all(ordered[i] == ordered[i - 1] + 1 for i in range(1, len(ordered)))


def suggest_chart_type(
    rows: list[list[str]],
    *,
    context_text: str = "",
) -> ChartType:
    """
    Smart Chart: line для временных рядов, bar для категорий/номенклатуры.

    * LINE — ключевые слова времени в headers/1-й колонке или последовательные годы.
    * BAR — товары, категории, ФИО и прочие номинальные подписи.
    """
    rows = normalize_table_rows(rows)
    if len(rows) < 2:
        return ChartType.BAR

    headers_lower = [h.strip().lower() for h in rows[0]]
    first_col_lower = [v.strip().lower() for v in _first_column_values(rows)]
    context_lower = (context_text or "").strip().lower()

    if _contains_time_axis_keyword(*headers_lower, *first_col_lower, context_lower):
        return ChartType.LINE

    if _is_strictly_sequential_integers(_first_column_values(rows)):
        return ChartType.LINE

    return ChartType.BAR


def _resolve_chart_type(
    rows: list[list[str]],
    chart_type: ChartType,
    *,
    context_text: str = "",
) -> ChartType:
    if chart_type is ChartType.AUTO:
        return suggest_chart_type(rows, context_text=context_text)
    return chart_type


def render_chart_png_bytes(
    rows: list[list[str]],
    chart_type: ChartType = ChartType.AUTO,
    *,
    context_text: str = "",
) -> tuple[bytes | None, ChartType]:
    wb_png = try_render_wb_chart_png(rows)
    if wb_png is not None:
        return wb_png, ChartType.BAR

    resolved = _resolve_chart_type(rows, chart_type, context_text=context_text)
    series = _extract_series(rows)
    if series is None:
        return None, resolved
    labels, values, header = series

    try:
        import matplotlib.pyplot as plt
    except ImportError:
        logger.warning("matplotlib not installed — chart skipped")
        return None, resolved

    fig, ax = plt.subplots(figsize=(8, 4.8), dpi=120)
    fig.patch.set_facecolor("#f8fafc")
    color = "#2563eb"

    if resolved is ChartType.PIE:
        ax.set_facecolor("#ffffff")
        wedges, _texts, autotexts = ax.pie(
            values,
            labels=labels,
            autopct="%1.1f%%",
            startangle=90,
            colors=plt.cm.Blues([0.35 + 0.5 * i / max(len(values), 1) for i in range(len(values))]),
        )
        for t in autotexts:
            t.set_fontsize(8)
        ax.set_title(header, fontsize=12, fontweight="bold", pad=12)
    else:
        ax.set_facecolor("#ffffff")
        if resolved is ChartType.LINE:
            ax.plot(labels, values, marker="o", linewidth=2.2, color=color)
        else:
            ax.bar(labels, values, color=color, alpha=0.88)
        ax.set_title(header, fontsize=12, fontweight="bold", pad=12)
        ax.grid(True, axis="y", linestyle="--", alpha=0.35)
        ax.tick_params(axis="x", rotation=25 if len(labels) > 4 else 0, labelsize=8)

    fig.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue(), resolved


def build_chart_png_bytes(
    rows: list[list[str]],
    *,
    context_text: str = "",
    chart_type: ChartType = ChartType.AUTO,
) -> bytes | None:
    png, _ = render_chart_png_bytes(rows, chart_type, context_text=context_text)
    return png


def build_xlsx_bytes(rows: list[list[str]]) -> tuple[bytes, float]:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    from services.table_number_parse import prepare_excel_value
    from services.table_text_response import compute_table_column_metrics

    rows = normalize_table_rows(rows)
    metrics = compute_table_column_metrics(rows)
    calculated_total = 0.0

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    header_font = Font(bold=True)
    total_fill = PatternFill(fill_type="solid", fgColor="E8F5E9")
    total_label_font = Font(bold=True)
    total_value_font = Font(bold=True, underline="double")
    money_number_format = "#,##0"

    if not rows:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue(), 0.0

    headers = metrics.headers if metrics else rows[0]
    data_rows = metrics.data_rows if metrics else rows[1:]
    value_col = metrics.value_col if metrics else -1
    label_col = metrics.label_col if metrics else 0
    add_total_row = metrics is not None and len(metrics.items) > 0

    ncols = max(len(headers), max((len(r) for r in data_rows), default=0))

    ws.append([prepare_excel_value(h) for h in headers])
    for c_idx in range(1, ncols + 1):
        ws.cell(row=1, column=c_idx).font = header_font

    for row in data_rows:
        padded = list(row) + [""] * max(0, ncols - len(row))
        excel_row: list[object] = []
        for c_idx, raw in enumerate(padded[:ncols]):
            prepared = prepare_excel_value(raw)
            excel_row.append(prepared)
            if c_idx == value_col and isinstance(prepared, (int, float)):
                calculated_total += float(prepared)
        ws.append(excel_row)
        row_idx = ws.max_row
        for c_idx in range(ncols):
            cell = ws.cell(row=row_idx, column=c_idx + 1)
            if isinstance(cell.value, (int, float)):
                cell.number_format = money_number_format

    if add_total_row and metrics is not None:
        last_data_row = 1 + len(data_rows)
        total_row_idx = last_data_row + 1
        if abs(calculated_total - round(calculated_total)) < 1e-9:
            total_cell_value: int | float = int(round(calculated_total))
        else:
            total_cell_value = round(calculated_total, 2)

        for c_idx in range(1, ncols + 1):
            cell = ws.cell(row=total_row_idx, column=c_idx)
            cell.fill = total_fill
            if c_idx == label_col + 1:
                cell.value = "Итого"
                cell.font = total_label_font
            elif c_idx == value_col + 1:
                cell.value = total_cell_value
                cell.font = total_value_font
                cell.number_format = money_number_format

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), calculated_total


def build_xlsx_bytes_from_table(headers: list[str], data_rows: list[list[str]]) -> tuple[bytes, float]:
    """Сборка Excel из ``headers`` + ``rows`` (без regex/Markdown)."""
    rows = normalize_table_rows([headers, *data_rows])
    return build_xlsx_bytes(rows)


def _build_pack_from_rows(
    rows: list[list[str]],
    *,
    title: str | None = None,
    context_text: str = "",
    chart_type: ChartType = ChartType.AUTO,
    telegram_rows: list[list[str]] | None = None,
    payload: TableJsonPayload | None = None,
    ai_insights: str | None = None,
    table_subrole: str | None = None,
) -> TableGeneratorPack | None:
    from services.table_subrole_types import normalize_table_subrole

    rows = normalize_table_rows(rows)
    if not rows:
        return None
    preview_rows = telegram_rows if telegram_rows else pick_telegram_preview_rows(rows)
    xlsx_bytes, calculated_total = build_xlsx_bytes(rows)
    subrole = normalize_table_subrole(table_subrole)
    caption: str | None = None

    if subrole == "wb_ozon_finance" and calculated_total > 0:
        wb_metrics = compute_wb_marketplace_metrics(rows, revenue_total=calculated_total)
        caption = build_wb_finance_express_html(
            calculated_total,
            wb_metrics=wb_metrics,
            matrix_rows=rows,
        )
    elif ai_insights is not None and payload is not None:
        caption = build_table_one_screen_html(
            payload,
            ai_insights=ai_insights,
            total_override=calculated_total,
            table_subrole=table_subrole,
        )
    elif payload is not None:
        caption = build_table_one_screen_html(
            payload,
            total_override=calculated_total,
            table_subrole=table_subrole,
        )
    if not caption and subrole == "wb_ozon_finance":
        caption = build_wb_telegram_preview_html(
            rows,
            title=title or "Отчёт NeuroMule",
            total_rub_override=calculated_total,
        )
    if not caption:
        caption = markdown_table_to_telegram_caption(preview_rows, title=title)
        if calculated_total > 0:
            from services.table_number_parse import format_rub_total

            caption = (
                f"💰 <b>ИТОГО:</b> {format_rub_total(calculated_total)}\n{caption}"
            )
    if len(caption) > _CAPTION_MAX:
        caption = caption[: _CAPTION_MAX - 1] + "…"
    chart_png, resolved = render_chart_png_bytes(rows, chart_type, context_text=context_text)
    return TableGeneratorPack(
        rows=rows,
        telegram_caption_html=caption,
        xlsx_bytes=xlsx_bytes,
        chart_png_bytes=chart_png,
        chart_type=resolved,
        calculated_total=calculated_total,
    )


def build_table_generator_pack(
    raw_json: str,
    *,
    context_text: str = "",
    chart_type: ChartType = ChartType.AUTO,
    ai_insights: str | None = None,
    table_subrole: str | None = None,
) -> TableGeneratorPack | None:
    """Собирает отчёт из JSON-ответа OpenRouter (роль table_generator)."""
    try:
        payload = parse_table_json_response(raw_json)
        if payload is None or not table_payload_has_data(payload):
            return None
        return _build_pack_from_rows(
            payload.to_rows_with_header(),
            title=payload.title,
            context_text=context_text,
            chart_type=chart_type,
            payload=payload,
            ai_insights=ai_insights,
            table_subrole=table_subrole,
        )
    except Exception:
        logger.exception("build_table_generator_pack failed")
        return None


def build_table_generator_pack_from_payload(
    payload: TableJsonPayload,
    *,
    context_text: str = "",
    chart_type: ChartType = ChartType.AUTO,
) -> TableGeneratorPack | None:
    return _build_pack_from_rows(
        payload.to_rows_with_header(),
        title=payload.title,
        context_text=context_text,
        chart_type=chart_type,
    )


def build_table_generator_pack_from_rows(
    rows: list[list[str]],
    *,
    context_text: str = "",
    chart_type: ChartType = ChartType.AUTO,
    title: str | None = None,
    preprocess: bool = True,
) -> TableGeneratorPack | None:
    """Локальная сборка из строк (например, после чтения .xlsx)."""
    display_title = title
    telegram_rows: list[list[str]] | None = None
    matrix = rows
    if preprocess:
        pre = preprocess_xlsx_rows(rows, title=title or "Отчёт NeuroMule")
        matrix = pre.rows
        display_title = pre.title
        telegram_rows = pre.telegram_rows
    return _build_pack_from_rows(
        matrix,
        title=display_title,
        context_text=context_text,
        chart_type=chart_type,
        telegram_rows=telegram_rows,
    )
