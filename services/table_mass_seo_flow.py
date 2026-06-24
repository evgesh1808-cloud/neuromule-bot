"""Массовая SEO-генерация для под-режима mass_seo_generation (до 20 строк)."""

from __future__ import annotations

import logging
from io import BytesIO

from config import Settings
from services import conversation as conv
from services import metrics
from services.ai_text import ask_ai_messages
from services.billing import billing
from services.billing.store import refund_charge
from services.dialog_write_worker import commit_assistant_turn_queued
from services.rate_limit_service import allow_request, rollback_last
from services.repository import dialog_append, insert_table_report
from services.table_number_parse import prepare_excel_value
from services.table_subrole_types import DEFAULT_TABLE_SUBROLE
from services.table_xlsx_flow import (
    DEEPSEEK_TABLE_FALLBACK_MODEL,
    TABLE_FALLBACK_TEMPERATURE,
    rows_to_canonical_table_json,
)
from services.table_xlsx_preprocess import preprocess_xlsx_rows
from services.use_cases.chat_turn import (
    ChatTurnOutcome,
    ChatTurnResult,
    _record_chat_success_billing,
)

logger = logging.getLogger(__name__)

_MAX_SEO_ROWS = 20
_SEO_COLUMN = "SEO_Описание_ИИ"
_NAME_HINTS = ("наименование", "название", "товар", "product", "title", "name", "артикул")


def _pick_name_column(headers: list[str]) -> int:
    lowered = [h.lower() for h in headers]
    for hint in _NAME_HINTS:
        for idx, header in enumerate(lowered):
            if hint in header:
                return idx
    return 0


def _build_seo_xlsx_bytes(rows: list[list[str]], seo_texts: list[str]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    matrix = rows
    if not matrix:
        wb = Workbook()
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(matrix[0])
    if _SEO_COLUMN not in headers:
        headers.append(_SEO_COLUMN)
    seo_col = headers.index(_SEO_COLUMN)

    wb = Workbook()
    ws = wb.active
    ws.title = "SEO"
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for i, row in enumerate(matrix[1:]):
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        if i < len(seo_texts):
            padded[seo_col] = seo_texts[i]
        ws.append([prepare_excel_value(c) for c in padded])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def _generate_seo_line(settings: Settings, product_name: str) -> str:
    prompt = (
        "Напиши короткое продающее SEO-описание товара для маркетплейса (2–4 предложения). "
        "Без HTML, без Markdown, только plain text.\n\n"
        f"Товар: {product_name.strip()}"
    )
    try:
        completion = await ask_ai_messages(
            settings,
            [
                {"role": "system", "content": "Ты — SEO-копирайтер маркетплейсов. Отвечай только текстом описания."},
                {"role": "user", "content": prompt},
            ],
            models=[DEEPSEEK_TABLE_FALLBACK_MODEL],
            max_tokens=400,
            temperature=TABLE_FALLBACK_TEMPERATURE,
            text_role="standard",
        )
        return (completion.get("content") or "").strip()
    except Exception:
        logger.debug("SEO line generation failed for %r", product_name[:80], exc_info=True)
        return ""


async def run_mass_seo_xlsx_turn(
    settings: Settings,
    user_id: int,
    rows: list[list[str]],
    *,
    file_name: str,
    title: str,
) -> ChatTurnResult:
    """Локальный Excel + до 20 SEO-описаний через DeepSeek (без JSON Mode)."""
    if not rows or not any(str(c).strip() for row in rows for c in row):
        return ChatTurnResult(outcome=ChatTurnOutcome.EMPTY_INPUT)

    if not await allow_request(settings, user_id, settings.chat_rate_limit_per_minute):
        return ChatTurnResult(outcome=ChatTurnOutcome.RATE_LIMITED)

    billing_result = await billing.resolve_and_charge_text_chat(user_id, "table_generator")
    plan = billing_result.plan
    charge_id = billing_result.charge_id
    effective_role = billing_result.effective_role_id

    if plan.blocked:
        await rollback_last(settings, user_id)
        if plan.block_reason in ("expert_role_requires_paid_tariff", "role_requires_smart_tariff"):
            return ChatTurnResult(outcome=ChatTurnOutcome.ROLE_NOT_ALLOWED)
        return ChatTurnResult(outcome=ChatTurnOutcome.INSUFFICIENT_BALANCE)

    pre = preprocess_xlsx_rows(rows, title=title)
    matrix = pre.rows
    if len(matrix) < 2:
        if charge_id:
            await refund_charge(charge_id)
        await rollback_last(settings, user_id)
        return ChatTurnResult(outcome=ChatTurnOutcome.AI_FAILED)

    headers = matrix[0]
    name_col = _pick_name_column(headers)
    data_rows = matrix[1 : 1 + _MAX_SEO_ROWS]

    seo_texts: list[str] = []
    for row in data_rows:
        name = (row[name_col] if name_col < len(row) else "").strip()
        if not name:
            seo_texts.append("")
            continue
        seo_texts.append(await _generate_seo_line(settings, name))

    xlsx_bytes = _build_seo_xlsx_bytes(matrix, seo_texts)
    table_json = rows_to_canonical_table_json(matrix, title=pre.title)
    if not table_json:
        if charge_id:
            await refund_charge(charge_id)
        await rollback_last(settings, user_id)
        return ChatTurnResult(outcome=ChatTurnOutcome.AI_FAILED)

    dialog_text = f"[📝 SEO Excel {file_name}]"
    await dialog_append(user_id, "user", dialog_text)
    await commit_assistant_turn_queued(user_id, table_json, settings.dialog_prune_keep)
    report_id = await insert_table_report(user_id, table_json)
    conv.schedule_memory_refresh(settings, user_id)
    metrics.incr("table.xlsx.mass_seo", {"outcome": "success"})
    _record_chat_success_billing(
        role=effective_role,
        energy_cost=plan.energy_cost,
        crystal_cost=plan.crystal_cost,
    )
    return ChatTurnResult(
        outcome=ChatTurnOutcome.SUCCESS,
        assistant_message=None,
        user_notice=billing_result.notice,
        effective_text_role=effective_role,
        table_raw_json=table_json,
        table_report_id=report_id,
        table_seo_xlsx_bytes=xlsx_bytes,
    )
