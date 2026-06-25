"""CPU-bound сборка table_generator в OS-thread pool (1 ядро — очередь 1)."""

from __future__ import annotations

import asyncio
import csv
import logging
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from services.table_chart_types import ChartType
from services.table_generator_pack import (
    TABLE_XLSX_FILENAME,
    _CAPTION_MAX,
    build_wb_telegram_preview_html,
    markdown_table_to_telegram_caption,
    render_chart_png_bytes,
)
from services.table_markdown import normalize_table_rows
from services.table_number_parse import format_rub_total, prepare_excel_value, safe_float
from services.table_subrole_types import DEFAULT_TABLE_SUBROLE, TableSubroleId, normalize_table_subrole
from services.table_text_response import (
    WbMarketplaceMetrics,
    build_table_one_screen_html,
    build_wb_finance_express_html,
    compute_table_column_metrics,
    compute_wb_marketplace_metrics,
)
from services.table_xlsx_preprocess import pick_telegram_preview_rows, preprocess_xlsx_rows
from services.file_processor import read_xlsx_rows_from_path

logger = logging.getLogger(__name__)

# Одно ядро CPU: тяжёлые таблицы строго по одной (async-очередь).
table_jobs_semaphore = asyncio.Semaphore(1)

_USN_RATE = 0.06
_WB_NET_PROFIT_FILL = "E2EFDA"
_ROI_NEGATIVE_FILL = "FFC7CE"

_TRAFFIC_COLUMN_HINTS: dict[str, tuple[str, ...]] = {
    "impressions": ("показ", "impression", "views"),
    "clicks": ("клик", "click"),
    "spend": ("расход", "затрат", "cost", "spend"),
    "leads": ("лид", "lead", "заявк"),
    "revenue": ("доход", "выруч", "revenue", "продаж"),
}


@dataclass(frozen=True)
class TableWorkerResult:
    """Результат синхронного воркера (без html_document)."""

    rows: list[list[str]]
    title: str
    telegram_caption_html: str
    xlsx_bytes: bytes
    chart_png_bytes: bytes | None
    chart_type: ChartType
    calculated_total: float
    xlsx_filename: str = TABLE_XLSX_FILENAME


def _read_csv_rows(path: Path, *, max_rows: int = 5000) -> list[list[str]]:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with path.open("r", encoding=encoding, newline="") as fh:
                reader = csv.reader(fh)
                rows: list[list[str]] = []
                for i, row in enumerate(reader):
                    if i >= max_rows:
                        break
                    cells = [str(c).strip() for c in row]
                    if any(cells):
                        rows.append(cells)
                if rows:
                    return rows
        except UnicodeDecodeError:
            continue
    return []


def _load_rows_from_path(file_path: str, *, is_csv: bool) -> list[list[str]]:
    path = Path(file_path)
    if is_csv:
        return _read_csv_rows(path)
    return read_xlsx_rows_from_path(path)


def _match_col(headers: list[str], hints: tuple[str, ...]) -> int | None:
    for idx, header in enumerate(headers):
        low = (header or "").lower()
        if any(h in low for h in hints):
            return idx
    return None


def _fmt_pct(value: float) -> str:
    return f"{value:.2f}%"


def _fmt_money(value: float) -> str:
    return f"{value:,.2f}".replace(",", " ")


def _append_finance_excel_rows(
    ws,
    *,
    ncols: int,
    label_col: int,
    value_col: int,
    calculated_total: float,
    wb_metrics: WbMarketplaceMetrics | None = None,
) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    tax = calculated_total * _USN_RATE
    net = calculated_total - tax
    net_fill = PatternFill(fill_type="solid", fgColor=_WB_NET_PROFIT_FILL)
    bold = Font(bold=True)
    money_fmt = "#,##0.00"
    pct_fmt = "0.00%"

    summary_rows: list[tuple[str, float, str, PatternFill | None]] = [
        ("Налог УСН 6%", tax, money_fmt, None),
        ("Чистая прибыль", net, money_fmt, net_fill),
    ]
    if wb_metrics is not None and wb_metrics.total_advertising_cost > 0:
        summary_rows.append(
            ("Удержания за продвижение", wb_metrics.total_advertising_cost, money_fmt, None)
        )
        summary_rows.append(
            ("Рекламная нагрузка", wb_metrics.ad_load_pct / 100.0, pct_fmt, None)
        )
    if wb_metrics is not None and wb_metrics.buyout_coef_pct > 0:
        summary_rows.append(
            ("Коэффициент выкупа", wb_metrics.buyout_coef_pct / 100.0, pct_fmt, None)
        )

    for label, amount, num_fmt, fill in summary_rows:
        row_idx = ws.max_row + 1
        for c_idx in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c_idx)
            if c_idx == label_col + 1:
                cell.value = label
                cell.font = bold
            elif c_idx == value_col + 1:
                cell.value = round(amount, 4) if num_fmt == pct_fmt else round(amount, 2)
                cell.font = bold
                cell.number_format = num_fmt
                if fill is not None:
                    cell.fill = fill

    for col_idx in range(1, ncols + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)


def _enrich_traffic_rows(rows: list[list[str]]) -> tuple[list[list[str]], list[int]]:
    """Добавляет CTR/CPC/CPA/ROI; возвращает индексы строк с ROI < 0 (data row offsets)."""
    matrix = normalize_table_rows(rows)
    if len(matrix) < 2:
        return matrix, []

    headers = list(matrix[0])
    imp_col = _match_col(headers, _TRAFFIC_COLUMN_HINTS["impressions"])
    clk_col = _match_col(headers, _TRAFFIC_COLUMN_HINTS["clicks"])
    spend_col = _match_col(headers, _TRAFFIC_COLUMN_HINTS["spend"])
    leads_col = _match_col(headers, _TRAFFIC_COLUMN_HINTS["leads"])
    rev_col = _match_col(headers, _TRAFFIC_COLUMN_HINTS["revenue"])

    extra_headers = ["CTR %", "CPC", "CPA", "ROI %"]
    if not any(h in headers for h in extra_headers):
        headers.extend(extra_headers)
    ctr_i = headers.index("CTR %")
    cpc_i = headers.index("CPC")
    cpa_i = headers.index("CPA")
    roi_i = headers.index("ROI %")

    out: list[list[str]] = [headers]
    negative_roi_rows: list[int] = []

    for row_idx, row in enumerate(matrix[1:], start=1):
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        impressions = safe_float(padded[imp_col]) if imp_col is not None else 0.0
        clicks = safe_float(padded[clk_col]) if clk_col is not None else 0.0
        spend = safe_float(padded[spend_col]) if spend_col is not None else 0.0
        leads = safe_float(padded[leads_col]) if leads_col is not None else 0.0
        revenue = safe_float(padded[rev_col]) if rev_col is not None else 0.0

        ctr = (clicks / impressions * 100.0) if impressions > 0 else 0.0
        cpc = (spend / clicks) if clicks > 0 else 0.0
        cpa = (spend / leads) if leads > 0 else 0.0
        roi = ((revenue - spend) / spend * 100.0) if spend > 0 else 0.0

        padded[ctr_i] = _fmt_pct(ctr)
        padded[cpc_i] = _fmt_money(cpc)
        padded[cpa_i] = _fmt_money(cpa)
        padded[roi_i] = _fmt_pct(roi)
        if roi < 0:
            negative_roi_rows.append(row_idx)
        out.append(padded)

    return out, negative_roi_rows


def _build_xlsx_for_subrole(
    rows: list[list[str]],
    subrole_id: TableSubroleId,
) -> tuple[bytes, float, list[int]]:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    negative_roi_rows: list[int] = []
    matrix = normalize_table_rows(rows)
    if subrole_id == "traffic_marketing":
        matrix, negative_roi_rows = _enrich_traffic_rows(matrix)

    metrics = compute_table_column_metrics(matrix)
    calculated_total = 0.0

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    header_font = Font(bold=True)
    total_fill = PatternFill(fill_type="solid", fgColor="E8F5E9")
    total_label_font = Font(bold=True)
    total_value_font = Font(bold=True, underline="double")
    roi_bad_fill = PatternFill(fill_type="solid", fgColor=_ROI_NEGATIVE_FILL)
    money_number_format = "#,##0"

    if not matrix:
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue(), 0.0, negative_roi_rows

    headers = metrics.headers if metrics else matrix[0]
    data_rows = metrics.data_rows if metrics else matrix[1:]
    value_col = metrics.value_col if metrics else -1
    label_col = metrics.label_col if metrics else 0
    ncols = max(len(headers), max((len(r) for r in data_rows), default=0))

    ws.append([prepare_excel_value(h) for h in headers])
    for c_idx in range(1, ncols + 1):
        ws.cell(row=1, column=c_idx).font = header_font

    roi_col_idx: int | None = None
    if subrole_id == "traffic_marketing" and "ROI %" in headers:
        roi_col_idx = headers.index("ROI %")

    for src_row_idx, row in enumerate(data_rows, start=1):
        padded = list(row) + [""] * max(0, ncols - len(row))
        excel_row: list[object] = []
        for c_idx, raw in enumerate(padded[:ncols]):
            prepared = prepare_excel_value(raw)
            excel_row.append(prepared)
            if c_idx == value_col and isinstance(prepared, (int, float)):
                calculated_total += float(prepared)
        ws.append(excel_row)
        row_idx = ws.max_row
        for c_idx in range(ncols):
            cell = ws.cell(row=row_idx, column=c_idx + 1)
            if isinstance(cell.value, (int, float)):
                cell.number_format = money_number_format
        if roi_col_idx is not None and src_row_idx in negative_roi_rows:
            cell = ws.cell(row=row_idx, column=roi_col_idx + 1)
            cell.fill = roi_bad_fill

    if metrics is not None and len(metrics.items) > 0:
        total_row_idx = ws.max_row + 1
        if abs(calculated_total - round(calculated_total)) < 1e-9:
            total_cell_value: int | float = int(round(calculated_total))
        else:
            total_cell_value = round(calculated_total, 2)
        for c_idx in range(1, ncols + 1):
            cell = ws.cell(row=total_row_idx, column=c_idx)
            cell.fill = total_fill
            if c_idx == label_col + 1:
                cell.value = "Итого"
                cell.font = total_label_font
            elif c_idx == value_col + 1:
                cell.value = total_cell_value
                cell.font = total_value_font
                cell.number_format = money_number_format

    if subrole_id == "wb_ozon_finance" and calculated_total > 0:
        wb_metrics = compute_wb_marketplace_metrics(matrix, revenue_total=calculated_total)
        _append_finance_excel_rows(
            ws,
            ncols=ncols,
            label_col=label_col,
            value_col=value_col,
            calculated_total=calculated_total,
            wb_metrics=wb_metrics,
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), calculated_total, negative_roi_rows


def _build_telegram_caption(
    rows: list[list[str]],
    *,
    title: str,
    subrole_id: TableSubroleId,
    calculated_total: float,
    telegram_rows: list[list[str]] | None,
    marketplace_platform: str | None = None,
) -> str:
    if subrole_id == "wb_ozon_finance" and calculated_total > 0:
        wb_metrics = compute_wb_marketplace_metrics(
            rows,
            revenue_total=calculated_total,
            platform=marketplace_platform,
        )
        caption = build_wb_finance_express_html(
            calculated_total,
            wb_metrics=wb_metrics,
            matrix_rows=rows,
            platform=marketplace_platform,
        )
        if len(caption) > _CAPTION_MAX:
            caption = caption[: _CAPTION_MAX - 1] + "…"
        return caption

    preview_rows = telegram_rows if telegram_rows else pick_telegram_preview_rows(rows)
    caption = ""
    if subrole_id == "wb_ozon_finance":
        caption = (
            build_wb_telegram_preview_html(
                rows,
                title=title,
                total_rub_override=calculated_total,
            )
            or ""
        )
    if not caption:
        matrix = normalize_table_rows(rows)
        if len(matrix) >= 2:
            import json

            from services.table_json import TableJsonPayload

            headers = [str(c) for c in matrix[0]]
            data_rows = [[str(c) for c in row] for row in matrix[1:]]
            payload = TableJsonPayload(
                title=title,
                headers=headers,
                rows=data_rows,
                raw_json=json.dumps(
                    {"title": title, "headers": headers, "rows": data_rows},
                    ensure_ascii=False,
                ),
            )
            caption = build_table_one_screen_html(
                payload,
                total_override=calculated_total if calculated_total > 0 else None,
                table_subrole=subrole_id,
            )
    if not caption:
        caption = markdown_table_to_telegram_caption(preview_rows, title=title)
        if calculated_total > 0:
            caption = f"💰 <b>ИТОГО:</b> {format_rub_total(calculated_total)}\n{caption}"

    if subrole_id == "traffic_marketing":
        caption = (
            "📈 <b>Рекламный трафик</b> — метрики CTR / CPC / CPA / ROI рассчитаны локально.\n"
            + caption
        )

    if len(caption) > _CAPTION_MAX:
        caption = caption[: _CAPTION_MAX - 1] + "…"
    return caption


def sync_table_processing_worker(
    file_path: str,
    subrole_id: str,
    is_csv: bool,
    *,
    title: str | None = None,
    marketplace_platform: str | None = None,
) -> TableWorkerResult | None:
    """
    Синхронный CPU-bound воркер: чтение файла, предобработка, Excel, график, caption.

    Вызывать только через :func:`run_table_processing_worker_async`.
    """
    sid = normalize_table_subrole(subrole_id)
    raw_rows = _load_rows_from_path(file_path, is_csv=is_csv)
    if not raw_rows:
        return None

    file_title = (title or Path(file_path).stem or "Отчёт NeuroMule").strip()
    pre = preprocess_xlsx_rows(raw_rows, title=file_title)
    matrix = pre.rows
    if not matrix:
        return None

    xlsx_bytes, calculated_total, _ = _build_xlsx_for_subrole(matrix, sid)
    if sid == "wb_ozon_finance" and pre.revenue_total > 0:
        calculated_total = pre.revenue_total
    caption = _build_telegram_caption(
        matrix,
        title=pre.title,
        subrole_id=sid,
        calculated_total=calculated_total,
        telegram_rows=pre.telegram_rows,
        marketplace_platform=marketplace_platform,
    )
    chart_png, resolved = render_chart_png_bytes(matrix, context_text=pre.title)

    return TableWorkerResult(
        rows=matrix,
        title=pre.title,
        telegram_caption_html=caption,
        xlsx_bytes=xlsx_bytes,
        chart_png_bytes=chart_png,
        chart_type=resolved,
        calculated_total=calculated_total,
    )


def sync_table_processing_from_rows(
    rows: list[list[str]],
    subrole_id: str,
    *,
    title: str = "Отчёт NeuroMule",
    marketplace_platform: str | None = None,
) -> TableWorkerResult | None:
    """Воркер без файла (уже загруженные строки)."""
    sid = normalize_table_subrole(subrole_id)
    pre = preprocess_xlsx_rows(rows, title=title)
    matrix = pre.rows
    if not matrix:
        return None

    xlsx_bytes, calculated_total, _ = _build_xlsx_for_subrole(matrix, sid)
    if sid == "wb_ozon_finance" and pre.revenue_total > 0:
        calculated_total = pre.revenue_total
    caption = _build_telegram_caption(
        matrix,
        title=pre.title,
        subrole_id=sid,
        calculated_total=calculated_total,
        telegram_rows=pre.telegram_rows,
        marketplace_platform=marketplace_platform,
    )
    chart_png, resolved = render_chart_png_bytes(matrix, context_text=pre.title)
    return TableWorkerResult(
        rows=matrix,
        title=pre.title,
        telegram_caption_html=caption,
        xlsx_bytes=xlsx_bytes,
        chart_png_bytes=chart_png,
        chart_type=resolved,
        calculated_total=calculated_total,
    )


async def run_table_processing_worker_async(
    file_path: str,
    subrole_id: str,
    is_csv: bool,
    *,
    title: str | None = None,
    marketplace_platform: str | None = None,
) -> TableWorkerResult | None:
    loop = asyncio.get_running_loop()
    async with table_jobs_semaphore:
        return await loop.run_in_executor(
            None,
            lambda: sync_table_processing_worker(
                file_path,
                subrole_id,
                is_csv,
                title=title,
                marketplace_platform=marketplace_platform,
            ),
        )


async def run_table_processing_from_rows_async(
    rows: list[list[str]],
    subrole_id: str,
    *,
    title: str = "Отчёт NeuroMule",
    marketplace_platform: str | None = None,
) -> TableWorkerResult | None:
    loop = asyncio.get_running_loop()
    async with table_jobs_semaphore:
        return await loop.run_in_executor(
            None,
            lambda: sync_table_processing_from_rows(
                rows,
                subrole_id,
                title=title,
                marketplace_platform=marketplace_platform,
            ),
        )
