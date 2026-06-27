"""ИИ-консалтинг для под-режима wb_ozon_finance (метрики ETL → OpenRouter)."""

from __future__ import annotations

import html
import json
import logging
import re
from dataclasses import dataclass, field
from typing import Any, Iterable

from config import Settings
from content.chat_prompt import build_wb_marketplace_finance_system_prompt
from services.ai_text import ask_ai_messages
from services.table_text_response import (
    WbMarketplaceMetrics,
    _fmt_rub_in_code,
    build_oos_forecast_line as _build_oos_forecast_line,
    build_oos_forecast_plain_summary as _build_oos_forecast_plain_summary,
    CFO_BUILD_FOOTER_HTML,
    CFO_BUILD_FOOTER_PLAIN,
    compute_wb_marketplace_metrics,
    FINANCE_REPORT_BUILD as _FINANCE_REPORT_BUILD,
    normalize_finance_report_html,
)
from services.telegram_safe_text import repair_telegram_html

logger = logging.getLogger(__name__)

_USN_RATE = 0.06
_BALLAST_BUYOUT_PCT = 15.0
_DRR_WARNING_PCT = 20.0
_LOW_BUYOUT_PCT = 40.0
_HIGH_DRR_PCT = 18.0
_CRITICAL_BUYOUT_PCT = 5.0
_SLOW_TURNOVER_DAYS = 45.0
_ILLIQUID_MIN_STOCK = 1.0
MIN_REVERSE_LOGISTICS_RUB = 50.0
_TOP_GROUP_A_DISPLAY = 5
_TOP_GROUP_C_DISPLAY = 5
_TOP_BALLAST_DISPLAY = 3
_TOP_NON_LIQUID_DISPLAY = 3
_TOP_LOSS_SKU_DISPLAY = 5
_TELEGRAM_MESSAGE_SOFT_MAX = 4000


def clamp_shop_returns_qty(
    returns_qty: float,
    *,
    sales_qty: float,
    deliveries_qty: float,
) -> float:
    """Возвраты на уровне магазина: не больше доставок и не больше 2× выкупов."""
    qty = max(0.0, returns_qty)
    if qty <= 0:
        return 0.0
    if deliveries_qty > 0:
        qty = min(qty, deliveries_qty)
    if sales_qty > 0:
        qty = min(qty, sales_qty * 2.0)
    return qty


def build_mpstats_return_logistics_payload(
    matrix_etl: object | None,
    *,
    revenue_total: float,
) -> dict[str, Any]:
    """Готовые строки потерь на обратной логистике для MPSTATS JSON и промпта."""
    if matrix_etl is None or revenue_total <= 0:
        return {
            "total_loss_rub": 0.0,
            "lines": [],
            "shop_avg_rub_per_unit": 0.0,
        }
    items: list[dict[str, Any]] = []
    for line in getattr(matrix_etl, "logistics_fomo_items", ()) or ():
        text = (line or "").strip()
        if text:
            items.append({"text": text})
    block = getattr(matrix_etl, "return_logistics_block", "") or ""
    lines = [ln.lstrip("• ").strip() for ln in block.splitlines() if ln.strip()]
    return {
        "total_loss_rub": round(float(getattr(matrix_etl, "logistics_fomo_rub", 0.0) or 0.0), 2),
        "lines": items or [{"text": ln} for ln in lines],
        "shop_avg_rub_per_unit": round(
            float(getattr(matrix_etl, "reverse_logistics_shop_avg", 0.0) or 0.0), 2
        ),
        "min_tariff_rub": MIN_REVERSE_LOGISTICS_RUB,
    }


def _format_sku_label(name: str, article: str) -> str:
    """Связка «название (арт. …)» — различает модификации одного товара."""
    name = (name or "—").strip()
    art = (article or "").strip()
    if not art or art == name:
        return name
    return f"{name} (арт. {art})"


def _sku_display_from_stats(sku_key: str, stats: dict[str, Any]) -> str:
    """Подпись SKU из CFO ``sku_data`` (human_name + артикул)."""
    name = (
        stats.get("human_name")
        or stats.get("short_name")
        or stats.get("name")
        or sku_key
    )
    return _format_sku_label(str(name), str(sku_key))


def _normalize_oos_stock_item(
    item: str | dict[str, Any],
    sku_data: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Приводит OOS-запись к dict с полным label «наименование (арт. SKU)»."""
    if isinstance(item, dict):
        out = dict(item)
        if not str(out.get("label") or "").strip():
            sku_key = str(out.get("article_id") or out.get("sku") or out.get("name") or "—")
            stats = (sku_data or {}).get(sku_key, {})
            name = out.get("name") or stats.get("human_name") or sku_key
            out["label"] = _format_sku_label(str(name), sku_key)
            out.setdefault("article_id", sku_key)
            out.setdefault("name", name)
        return out
    sku_key = str(item).strip()
    stats = (sku_data or {}).get(sku_key, {})
    name = stats.get("human_name") or stats.get("short_name") or sku_key
    return {
        "name": str(name),
        "article_id": sku_key,
        "sku": sku_key,
        "label": _sku_display_from_stats(sku_key, stats),
        "stock_qty": 0,
    }


def _oos_item_display_label(item: dict[str, Any]) -> str:
    """Полная подпись для OOS/ABC/плана — всегда с артикулом при необходимости."""
    label = str(item.get("label") or "").strip()
    if label:
        return label
    return _format_sku_label(
        str(item.get("name") or "—"),
        str(item.get("article_id") or item.get("sku") or "—"),
    )


def _sku_qualifies_for_success_zone(buyout_pct: float, margin_rub: float) -> bool:
    """🟢 Зона успеха: выкуп ≥ порога, маржа > 0, выкуп не нулевой."""
    return (
        buyout_pct >= _LOW_BUYOUT_PCT
        and buyout_pct > 0
        and margin_rub > 0
    )


def _dedupe_report_noise(text: str) -> str:
    """Убирает дубли вроде «(риск OOS) (риск OOS)»."""
    cleaned = (text or "").strip()
    cleaned = re.sub(r"(\(риск OOS\)\s*)+", "(риск OOS) ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    return cleaned.strip()


def _sku_label_is_valid(name: str, article: str) -> bool:
    """Строка с пустым прочерком вместо имени не выводится в отчёт."""
    from services.wb_transaction_parse import is_valid_wb_sku

    return is_valid_wb_sku(name, article)


def _ru_more_goods_suffix(count: int) -> str:
    """«… и ещё N товара/товаров» с правильным склонением."""
    n = abs(int(count))
    if n % 10 == 1 and n % 100 != 11:
        word = "товар"
    elif n % 10 in (2, 3, 4) and n % 100 not in (12, 13, 14):
        word = "товара"
    else:
        word = "товаров"
    return f"… и ещё {n} {word}"


def _strict_non_liquid(*, revenue: float, stock_qty: float) -> bool:
    """Неликвид: остаток на складе есть, выручка за период — ноль."""
    return stock_qty > 0 and revenue == 0.0


def _classify_group_c_problem_zone(
    buyout_pct: float,
    revenue: float,
    *,
    stock_qty: float,
    sales_qty: float,
    days_until_stockout: float | None,
) -> str:
    """«ballast» — низкий выкуп; «illiquid» — залежалый остаток без спроса."""
    if buyout_pct < _BALLAST_BUYOUT_PCT and sales_qty > 0:
        return "ballast"
    if sales_qty <= 0 and stock_qty >= _ILLIQUID_MIN_STOCK:
        return "illiquid"
    if (
        revenue <= 0
        and stock_qty >= _ILLIQUID_MIN_STOCK
        and buyout_pct <= _CRITICAL_BUYOUT_PCT
    ):
        return "illiquid"
    if (
        days_until_stockout is not None
        and days_until_stockout > _SLOW_TURNOVER_DAYS
        and sales_qty > 0
        and stock_qty >= _ILLIQUID_MIN_STOCK
    ):
        return "illiquid"
    if buyout_pct < _BALLAST_BUYOUT_PCT:
        return "ballast"
    if revenue <= 0 and buyout_pct < 55.0:
        return "ballast"
    return "ballast" if buyout_pct < 55.0 else "illiquid"


def _parse_return_logistics_from_etl(matrix_etl: object | None) -> dict[str, dict[str, float | int]]:
    """Парсит готовые строки ETL «Логистика возвратов: …» в словарь по метке SKU."""
    out: dict[str, dict[str, float | int]] = {}
    if matrix_etl is None:
        return out
    pattern = re.compile(
        r"Логистика возвратов:\s*(.+?):\s*(\d+)\s*возвратов\.\s*"
        r"Общий убыток на пустых покатушках:\s*≈\s*([\d\s.,]+)\s*руб",
        re.IGNORECASE,
    )
    block = getattr(matrix_etl, "return_logistics_block", "") or ""
    for line in block.splitlines():
        m = pattern.search(line)
        if not m:
            continue
        label = m.group(1).strip()
        loss_raw = m.group(3).replace(" ", "").replace(",", ".")
        try:
            loss = float(loss_raw)
        except ValueError:
            loss = 0.0
        out[label] = {
            "returns": int(m.group(2)),
            "loss": round(loss, 2),
        }
    return out


def _extract_problem_zones_structured(
    matrix_etl: object | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Полные списки балласта и неликвида для JSON (без усечения)."""
    if matrix_etl is None:
        return [], []
    group_c = getattr(matrix_etl, "abc_group_c", ()) or ()
    if not group_c:
        return [], []

    logistics_map = _parse_return_logistics_from_etl(matrix_etl)
    oos_map = {f.label: f for f in getattr(matrix_etl, "oos_forecasts", ()) or ()}
    catalog = {s.name: s for s in getattr(matrix_etl, "sku_catalog", ()) or ()}
    ballast: list[dict[str, Any]] = []
    non_liquid: list[dict[str, Any]] = []

    for sku in group_c:
        oos = oos_map.get(sku.name)
        stock = float(oos.stock_qty) if oos else 0.0
        sales = float(oos.sales_period_qty) if oos else 0.0
        days = oos.days_until_stockout if oos else None
        label = _format_sku_label(sku.name, sku.article_id)
        if _strict_non_liquid(revenue=sku.revenue, stock_qty=stock):
            detail = catalog.get(sku.name)
            unit_cost = float(detail.unit_cost_rub) if detail else 0.0
            frozen = round(stock * unit_cost, 2) if unit_cost > 0 else 0.0
            non_liquid.append(
                {
                    "sku": label,
                    "name": sku.name,
                    "article_id": sku.article_id,
                    "stock": int(stock),
                    "revenue": 0.0,
                    "cost": round(unit_cost, 2),
                    "frozen_capital_rub": frozen,
                }
            )
        elif _classify_group_c_problem_zone(
            sku.buyout_pct,
            sku.revenue,
            stock_qty=stock,
            sales_qty=sales,
            days_until_stockout=days,
        ) == "ballast" or (
            sku.buyout_pct < _BALLAST_BUYOUT_PCT and sales > 0
        ):
            log = logistics_map.get(label) or logistics_map.get(sku.name) or {}
            ballast.append(
                {
                    "sku": label,
                    "buyout": round(float(sku.buyout_pct), 1),
                    "returns": int(log.get("returns", 0)),
                    "loss": round(float(log.get("loss", 0.0)), 2),
                }
            )

    ballast.sort(key=lambda x: float(x.get("buyout", 100.0)))
    non_liquid.sort(key=lambda x: -int(x.get("stock", 0)))
    return ballast, non_liquid


def _sku_unit_profit_rub(net_profit: float, sales_qty: float) -> float:
    if sales_qty > 0:
        return round(net_profit / sales_qty, 2)
    return round(net_profit, 2)


def _collect_etl_dynamic_slices(
    matrix_etl: object | None,
) -> tuple[
    tuple[dict[str, Any], ...],
    tuple[dict[str, Any], ...],
    tuple[dict[str, Any], ...],
    tuple[dict[str, Any], ...],
    float,
]:
    """Срезы ETL для светофора, калькулятора потерь и плана действий."""
    if matrix_etl is None:
        return (), (), (), (), 0.0

    oos_map = {f.label: f for f in getattr(matrix_etl, "oos_forecasts", ()) or ()}
    catalog_map = {s.name: s for s in getattr(matrix_etl, "sku_catalog", ()) or ()}

    group_a_items: list[dict[str, Any]] = []
    for sku in getattr(matrix_etl, "abc_group_a", ()) or ():
        if not _sku_label_is_valid(sku.name, sku.article_id):
            continue
        detail = catalog_map.get(sku.name)
        oos = oos_map.get(sku.name)
        sales_qty = float(detail.sales_qty) if detail and detail.sales_qty > 0 else 0.0
        if sales_qty <= 0 and oos:
            sales_qty = float(oos.sales_period_qty)
        group_a_items.append(
            {
                "name": sku.name,
                "article_id": sku.article_id,
                "label": _format_sku_label(sku.name, sku.article_id),
                "unit_profit_rub": _sku_unit_profit_rub(sku.net_profit, sales_qty),
                "net_profit_rub": round(float(sku.net_profit), 2),
                "buyout_pct": round(float(sku.buyout_pct), 1),
            }
        )
    group_a_items.sort(key=lambda x: float(x["net_profit_rub"]), reverse=True)

    loss_items: list[dict[str, Any]] = []
    for detail in getattr(matrix_etl, "sku_catalog", ()) or ():
        if not _sku_label_is_valid(detail.name, detail.article_id):
            continue
        if detail.net_profit >= 0:
            continue
        oos = oos_map.get(detail.name)
        sales_qty = float(detail.sales_qty) if detail.sales_qty > 0 else 0.0
        if sales_qty <= 0 and oos:
            sales_qty = float(oos.sales_period_qty)
        loss_items.append(
            {
                "name": detail.name,
                "article_id": detail.article_id,
                "label": _format_sku_label(detail.name, detail.article_id),
                "net_profit_rub": round(float(detail.net_profit), 2),
                "unit_profit_rub": _sku_unit_profit_rub(detail.net_profit, sales_qty),
            }
        )
    loss_items.sort(key=lambda x: float(x["net_profit_rub"]))

    _, non_liquid = _extract_problem_zones_structured(matrix_etl)
    non_liquid_items = tuple(non_liquid)
    frozen_total = round(
        sum(float(item.get("frozen_capital_rub", 0.0)) for item in non_liquid_items),
        2,
    )

    oos_zero: list[dict[str, Any]] = []
    oos_critical: list[dict[str, Any]] = []
    zero_keys: set[tuple[str, str]] = set()
    for forecast in getattr(matrix_etl, "oos_forecasts", ()) or ():
        stock = float(forecast.stock_qty)
        sales = float(forecast.sales_period_qty)
        detail = catalog_map.get(forecast.label)
        article_id = detail.article_id if detail else forecast.label
        key = (forecast.label, article_id)
        if stock <= 0:
            zero_keys.add(key)
            oos_zero.append(
                {
                    "name": forecast.label,
                    "article_id": article_id,
                    "label": _format_sku_label(forecast.label, article_id),
                    "stock_qty": 0,
                    "sales_qty": round(sales, 2),
                }
            )
            continue
        days = forecast.days_until_stockout
        if sales <= 0 or not getattr(forecast, "risk_out_of_stock", False):
            continue
        if days is None or float(days) > 5.0:
            continue
        if key in zero_keys:
            continue
        oos_critical.append(
            {
                "name": forecast.label,
                "article_id": article_id,
                "label": _format_sku_label(forecast.label, article_id),
                "stock_qty": round(stock, 2),
                "days_until_stockout": round(float(days), 1),
            }
        )
    oos_critical.sort(key=lambda x: float(x.get("days_until_stockout") or 999.0))

    return (
        tuple(group_a_items),
        tuple(loss_items),
        non_liquid_items,
        tuple(oos_zero),
        tuple(oos_critical),
        frozen_total,
    )


def _escape_oos_html(text: str) -> str:
    import html as html_module

    return html_module.escape(text or "", quote=False)


def _wrap_finance_report_in_pre(html: str) -> str:
    """Моноширинная сетка разделителей для Telegram HTML."""
    inner = (html or "").strip()
    if not inner:
        return ""
    if inner.startswith("<pre>") and inner.endswith("</pre>"):
        return inner
    return f"<pre>{inner}</pre>"


@dataclass(frozen=True)
class WbFinanceMatrixAggregation:
    """Схлопнутые списки матрицы для Telegram (математика — по всем SKU)."""

    abc_a_display_lines: tuple[str, ...] = ()
    tail_a_count: int = 0
    abc_c_display_lines: tuple[str, ...] = ()
    tail_c_count: int = 0
    tail_c_revenue: float = 0.0
    ballast_display_lines: tuple[str, ...] = ()
    tail_ballast_count: int = 0
    tail_ballast_loss: float = 0.0
    non_liquid_display_lines: tuple[str, ...] = ()
    tail_frozen_count: int = 0
    tail_frozen_stock: int = 0
    loss_sku_display_lines: tuple[str, ...] = ()
    tail_loss_sku_count: int = 0
    tail_loss_sku_rub: float = 0.0
    abc_group_a_display_items: tuple[dict[str, Any], ...] = ()
    non_liquid_display_items: tuple[dict[str, Any], ...] = ()


def _escape_verdict(text: str) -> str:
    from services.telegram_safe_text import _escape_telegram_html

    return _escape_telegram_html(text)


def _tail_line_group_a(count: int) -> str:
    return f"• <i>...и ещё {count} успешных товаров группы А</i>"


def _tail_line_group_c(count: int, revenue: float) -> str:
    return (
        f"• <i>...и ещё {count} товаров с низким спросом "
        f"(суммарная выручка: {_fmt_rub_in_code(revenue)} руб.)</i>"
    )


def _tail_line_ballast(count: int, loss: float) -> str:
    return (
        f"• <i>Оставшиеся {count} товаров-балласта принесли убыток на покатушках: "
        f"{_fmt_rub_in_code(loss)} руб.</i>"
    )


def _tail_line_non_liquid(count: int, stock: int) -> str:
    return (
        f"• <i>Оставшиеся {count} позиций неликвида заморозили на складе "
        f"{stock} шт. товара.</i>"
    )


def _tail_line_loss_skus(count: int, loss_rub: float) -> str:
    return (
        f"• <i>...и ещё {count} убыточных SKU на сумму "
        f"{_fmt_rub_in_code(abs(loss_rub))} руб.</i>"
    )


def aggregate_matrix_display(matrix_etl: object | None) -> WbFinanceMatrixAggregation:
    """ТОП-N в тексте, хвост — агрегированными метриками (все SKU учтены в суммах)."""
    empty = WbFinanceMatrixAggregation()
    if matrix_etl is None:
        return empty

    oos_map = {f.label: f for f in getattr(matrix_etl, "oos_forecasts", ()) or ()}
    catalog_map = {s.name: s for s in getattr(matrix_etl, "sku_catalog", ()) or ()}

    group_a_sorted = sorted(
        getattr(matrix_etl, "abc_group_a", ()) or (),
        key=lambda s: float(s.net_profit),
        reverse=True,
    )
    abc_a_lines: list[str] = []
    for sku in group_a_sorted[:_TOP_GROUP_A_DISPLAY]:
        if not _sku_label_is_valid(sku.name, sku.article_id):
            continue
        detail = catalog_map.get(sku.name)
        oos = oos_map.get(sku.name)
        sales_qty = float(detail.sales_qty) if detail and detail.sales_qty > 0 else 0.0
        if sales_qty <= 0 and oos:
            sales_qty = float(oos.sales_period_qty)
        unit_profit = _sku_unit_profit_rub(sku.net_profit, sales_qty)
        label = _format_sku_label(sku.name, sku.article_id)
        safe_label = _escape_verdict(label)
        abc_a_lines.append(
            f"• <b>{safe_label}</b> — чистая прибыль "
            f"<code>{_fmt_rub_in_code(unit_profit)}</code>/шт., "
            f"выкуп <code>{sku.buyout_pct:.1f}%</code>"
        )
    tail_a_count = max(0, len(group_a_sorted) - _TOP_GROUP_A_DISPLAY)

    group_c_sorted = sorted(
        getattr(matrix_etl, "abc_group_c", ()) or (),
        key=lambda s: float(s.revenue),
    )
    abc_c_lines: list[str] = []
    for sku in group_c_sorted[:_TOP_GROUP_C_DISPLAY]:
        if not _sku_label_is_valid(sku.name, sku.article_id):
            continue
        label = _format_sku_label(sku.name, sku.article_id)
        abc_c_lines.append(
            f"• {_escape_verdict(label)} — выручка "
            f"<code>{_fmt_rub_in_code(sku.revenue)}</code> руб."
        )
    tail_c = group_c_sorted[_TOP_GROUP_C_DISPLAY:]
    tail_c_count = len(tail_c)
    tail_c_revenue = round(sum(float(s.revenue) for s in tail_c), 2)

    ballast, non_liquid = _extract_problem_zones_structured(matrix_etl)
    ballast_lines: list[str] = []
    for item in ballast[:_TOP_BALLAST_DISPLAY]:
        ballast_lines.append(
            f"• {_escape_verdict(str(item['sku']))} — выкуп "
            f"<code>{item['buyout']:.1f}%</code>, убыток на покатушках "
            f"<code>{_fmt_rub_in_code(float(item['loss']))}</code> руб."
        )
    tail_ballast = ballast[_TOP_BALLAST_DISPLAY:]
    tail_ballast_count = len(tail_ballast)
    tail_ballast_loss = round(sum(float(i.get("loss", 0.0)) for i in tail_ballast), 2)

    non_liquid_lines: list[str] = []
    for item in non_liquid[:_TOP_NON_LIQUID_DISPLAY]:
        stock = int(item.get("stock", 0))
        non_liquid_lines.append(
            f"• {_escape_verdict(str(item['sku']))} — остаток <code>{stock}</code> шт."
        )
    tail_frozen = non_liquid[_TOP_NON_LIQUID_DISPLAY:]
    tail_frozen_count = len(tail_frozen)
    tail_frozen_stock = sum(int(i.get("stock", 0)) for i in tail_frozen)

    loss_items: list[dict[str, Any]] = []
    for detail in getattr(matrix_etl, "sku_catalog", ()) or ():
        if detail.net_profit >= 0:
            continue
        loss_items.append(
            {
                "label": _format_sku_label(detail.name, detail.article_id),
                "net_profit_rub": round(float(detail.net_profit), 2),
            }
        )
    loss_items.sort(key=lambda x: float(x["net_profit_rub"]))
    loss_lines: list[str] = []
    for item in loss_items[:_TOP_LOSS_SKU_DISPLAY]:
        loss_lines.append(
            f"• <b>{_escape_verdict(str(item['label']))}</b> — убыток "
            f"<code>{_fmt_rub_in_code(abs(float(item['net_profit_rub'])))}</code> руб."
        )
    tail_loss = loss_items[_TOP_LOSS_SKU_DISPLAY:]
    tail_loss_count = len(tail_loss)
    tail_loss_rub = round(sum(float(i["net_profit_rub"]) for i in tail_loss), 2)

    group_a_display_items: list[dict[str, Any]] = []
    for sku in group_a_sorted[:_TOP_GROUP_A_DISPLAY]:
        detail = catalog_map.get(sku.name)
        oos = oos_map.get(sku.name)
        sales_qty = float(detail.sales_qty) if detail and detail.sales_qty > 0 else 0.0
        if sales_qty <= 0 and oos:
            sales_qty = float(oos.sales_period_qty)
        group_a_display_items.append(
            {
                "name": sku.name,
                "article_id": sku.article_id,
                "label": _format_sku_label(sku.name, sku.article_id),
                "unit_profit_rub": _sku_unit_profit_rub(sku.net_profit, sales_qty),
                "net_profit_rub": round(float(sku.net_profit), 2),
                "buyout_pct": round(float(sku.buyout_pct), 1),
            }
        )

    return WbFinanceMatrixAggregation(
        abc_a_display_lines=tuple(abc_a_lines),
        tail_a_count=tail_a_count,
        abc_c_display_lines=tuple(abc_c_lines),
        tail_c_count=tail_c_count,
        tail_c_revenue=tail_c_revenue,
        ballast_display_lines=tuple(ballast_lines),
        tail_ballast_count=tail_ballast_count,
        tail_ballast_loss=tail_ballast_loss,
        non_liquid_display_lines=tuple(non_liquid_lines),
        tail_frozen_count=tail_frozen_count,
        tail_frozen_stock=tail_frozen_stock,
        loss_sku_display_lines=tuple(loss_lines),
        tail_loss_sku_count=tail_loss_count,
        tail_loss_sku_rub=tail_loss_rub,
        abc_group_a_display_items=tuple(group_a_display_items),
        non_liquid_display_items=tuple(non_liquid[:_TOP_NON_LIQUID_DISPLAY]),
    )


def build_matrix_problem_zones_block_from_aggregation(
    aggregation: WbFinanceMatrixAggregation,
) -> str:
    """Балласт и неликвид с агрегированным хвостом."""
    lines: list[str] = []
    if aggregation.ballast_display_lines:
        lines.append("📉 <b>Балласт (Деньги уходят на пустые покатушки):</b>")
        lines.extend(aggregation.ballast_display_lines)
        if aggregation.tail_ballast_count > 0:
            lines.append(
                _tail_line_ballast(
                    aggregation.tail_ballast_count,
                    aggregation.tail_ballast_loss,
                )
            )
    if aggregation.non_liquid_display_lines:
        lines.append("❄️ <b>Неликвид (Капитал заморожен на складе):</b>")
        lines.extend(aggregation.non_liquid_display_lines)
        if aggregation.tail_frozen_count > 0:
            lines.append(
                _tail_line_non_liquid(
                    aggregation.tail_frozen_count,
                    aggregation.tail_frozen_stock,
                )
            )
    return "\n".join(lines) if lines else "проблемных зон в группе C не выявлено"


def _extract_problem_zone_sku_lists(matrix_etl: object | None) -> tuple[list[str], list[str]]:
    """Устаревшая обёртка: метки SKU из структурированных зон."""
    ballast, non_liquid = _extract_problem_zones_structured(matrix_etl)
    return (
        [str(item["sku"]) for item in ballast],
        [str(item["sku"]) for item in non_liquid],
    )


def _build_traffic_light_json(
    *,
    group_a: list[str],
    ballast: list[dict[str, Any]],
    drr: float,
    prompt_metrics: WbFinancePromptMetrics | None,
    wb_metrics: WbMarketplaceMetrics | None,
) -> dict[str, str]:
    """Готовые тексты светофора — модель только копирует в HTML."""
    green_parts: list[str] = []
    yellow_parts: list[str] = []
    red_parts: list[str] = []

    if group_a:
        green_parts.append("Прибыльные лидеры: " + "; ".join(group_a))
    if prompt_metrics and prompt_metrics.abc_a_leader_name:
        green_parts.append(
            f"Лидер «{prompt_metrics.abc_a_leader_name}» "
            f"(арт. {prompt_metrics.abc_a_leader_article})"
        )

    if drr > _DRR_WARNING_PCT:
        yellow_parts.append(
            f"ДРР {drr:.1f}% — Вы работаете на рекламу, а не на карман."
        )
    elif drr > 0:
        yellow_parts.append(f"ДРР {drr:.1f}% — контролируйте окупаемость рекламы.")

    if prompt_metrics and prompt_metrics.buy_ratio_pct > 0 and prompt_metrics.buy_ratio_pct < 55:
        yellow_parts.append(f"Выкуп {prompt_metrics.buy_ratio_pct:.1f}% — зона внимания.")

    for item in ballast:
        red_parts.append(
            f"«{item['sku']}» — выкуп {item['buyout']}% — "
            f"убыток на покатушках ≈ {item['loss']} руб."
        )
    if prompt_metrics and prompt_metrics.outsider_loss > 0:
        red_parts.append(
            f"«{prompt_metrics.outsider_name}» (арт. {prompt_metrics.outsider_article}) — "
            f"убыток {_fmt_rub_in_code(prompt_metrics.outsider_loss)} руб."
        )
    if wb_metrics and wb_metrics.top5_units:
        worst = min(wb_metrics.top5_units, key=lambda u: u.net_income)
        if worst.net_income < 0:
            red_parts.append(
                f"«{worst.label}» — {_fmt_rub_in_code(worst.net_income)} руб./шт."
            )

    return {
        "green": " ".join(green_parts) if green_parts else "критических утечек в лидерах не выявлено",
        "yellow": " ".join(yellow_parts) if yellow_parts else "показатели в норме",
        "red": " ".join(red_parts) if red_parts else "критических убытков не выявлено",
    }


def _build_oos_predictions_map(matrix_etl: object | None, *, max_days: int = 7) -> dict[str, int]:
    """Критические OOS: дней до обнуления остатка (≤ max_days)."""
    if matrix_etl is None:
        return {}
    catalog = {s.name: s for s in getattr(matrix_etl, "sku_catalog", ()) or ()}
    out: dict[str, int] = {}
    for forecast in getattr(matrix_etl, "oos_forecasts", ()) or ():
        days = forecast.days_until_stockout
        if days is None or days > max_days:
            continue
        if forecast.stock_qty <= 0 and forecast.sales_period_qty <= 0:
            continue
        detail = catalog.get(forecast.label)
        key = (
            _format_sku_label(detail.name, detail.article_id)
            if detail
            else forecast.label
        )
        out[key] = max(0, int(days))
    return out


_WB_XLSX_HEADER_SCAN_ROWS = 25
_WB_RETURN_COL_ID_SKIP = (
    "srid",
    "rrid",
    "rrd",
    " id",
    "id ",
    "номер",
    "код возврата",
    "документ",
    "транзак",
)
_WB_QTY_MARKERS = ("шт", "кол-во", "количество", "единиц")


def _wb_xlsx_is_return_id_header(header: str) -> bool:
    low = (header or "").lower()
    if any(q in low for q in _WB_QTY_MARKERS):
        return False
    return any(s in low for s in _WB_RETURN_COL_ID_SKIP)


def _wb_xlsx_find_col(
    headers: list[str],
    keywords: tuple[str, ...],
    *,
    require_qty: bool = False,
    skip_return_ids: bool = False,
) -> int | None:
    for idx, header in enumerate(headers):
        low = (header or "").lower()
        if not any(k.lower() in low for k in keywords):
            continue
        if require_qty and not any(q in low for q in _WB_QTY_MARKERS):
            continue
        if skip_return_ids and _wb_xlsx_is_return_id_header(header):
            continue
        return idx
    return None


def _wb_xlsx_safe_float(value: object) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _wb_xlsx_sku_label(brand: str, article: str) -> str:
    brand = (brand or "").strip()
    article = (article or "").strip()
    if not brand or brand in ("—", "-", "Unknown", "None"):
        brand = article or "—"
    if not article or article == brand:
        return brand
    return f"{brand} (арт. {article})"


def _wb_xlsx_skip_row(brand: str, article: str) -> bool:
    for label in (brand, article):
        low = (label or "").strip().lower()
        if low.startswith(("итого", "всего", "total")):
            return True
    if brand in ("None", "", "—", "-") and article in ("None", "", "—", "-"):
        return True
    return False


def _wb_xlsx_detect_header(sheet: object) -> tuple[int, list[str]] | None:
    """Ищет строку шапки WB (после преамбули поставщика)."""
    max_row = min(getattr(sheet, "max_row", 0) or 0, _WB_XLSX_HEADER_SCAN_ROWS)
    max_col = min(getattr(sheet, "max_column", 0) or 0, 80)
    revenue_keys = (
        "к перечислению",
        "вайлдберриз к перечислению",
        "продажи",
        "выруч",
    )
    identity_keys = ("бренд", "артикул", "предмет", "наименование", "номенклатур")
    for row_idx in range(1, max_row + 1):
        headers = [
            str(sheet.cell(row=row_idx, column=col).value or "").strip()
            for col in range(1, max_col + 1)
        ]
        if not any(headers):
            continue
        has_revenue = any(any(k in (h or "").lower() for k in revenue_keys) for h in headers)
        has_identity = any(any(k in (h or "").lower() for k in identity_keys) for h in headers)
        if has_revenue and has_identity:
            return row_idx, headers
    return None


def _wb_xlsx_return_loss_rub(returns_count: int, delivery_rub: float) -> float:
    if returns_count <= 0:
        return 0.0
    per_unit = delivery_rub / returns_count if delivery_rub > 0 else 0.0
    unit = max(MIN_REVERSE_LOGISTICS_RUB, per_unit)
    return round(returns_count * unit, 2)


def build_wb_weekly_xlsx_ai_context(file_path: str | Path) -> dict[str, Any] | None:
    """
    Парсер еженедельных отчётов Wildberries (openpyxl): агрегация по SKU,
    валидация возвратов, MPSTATS JSON без галлюцинаций модели.
    """
    from pathlib import Path as PathCls

    from openpyxl import load_workbook

    path = PathCls(file_path)
    if not path.is_file():
        return None

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        detected = _wb_xlsx_detect_header(sheet)
        if detected is None:
            return None
        header_row, headers = detected

        col_brand = _wb_xlsx_find_col(headers, ("бренд", "brand"))
        col_art = _wb_xlsx_find_col(
            headers, ("артикул поставщика", "артикул", "vendor", "sku", "barcode")
        )
        col_name = _wb_xlsx_find_col(headers, ("предмет", "наименование", "номенклатур", "товар"))
        col_sales = _wb_xlsx_find_col(
            headers,
            (
                "вайлдберриз к перечислению",
                "к перечислению за товар",
                "к перечислению",
                "продажи",
                "выруч",
            ),
        )
        col_cost = _wb_xlsx_find_col(headers, ("себестоимость", "закупка"))
        col_delivery = _wb_xlsx_find_col(
            headers,
            ("услуги по доставке", "логистика", "стоимость логистики", "доставк"),
        )
        col_returns_qty = _wb_xlsx_find_col(
            headers,
            ("количество возвратов", "возврат штук", "возвраты"),
            require_qty=True,
            skip_return_ids=True,
        )
        if col_returns_qty is None:
            col_returns_qty = _wb_xlsx_find_col(
                headers, ("возврат",), require_qty=True, skip_return_ids=True
            )
        col_orders_qty = _wb_xlsx_find_col(
            headers,
            ("количество заказов", "заказы штук", "заказы", "заказано"),
            require_qty=True,
        )
        if col_orders_qty is None:
            col_orders_qty = _wb_xlsx_find_col(
                headers, ("доставк", "выкупили"), require_qty=True
            )
        col_ad_spend = _wb_xlsx_find_col(
            headers, ("расходы на рекламу", "реклама", "внутренняя реклама", "продвижен", "удержан")
        )
        col_stock = _wb_xlsx_find_col(headers, ("текущий остаток", "остаток на складе", "остатк"))
        col_daily_sales = _wb_xlsx_find_col(
            headers, ("средние продажи в день", "скорость продаж")
        )

        if col_sales is None:
            return None

        def cell(row_idx: int, col_idx: int | None) -> object:
            if col_idx is None:
                return None
            return sheet.cell(row=row_idx, column=col_idx + 1).value

        sku_data: dict[str, dict[str, float | int]] = {}
        max_row = getattr(sheet, "max_row", header_row) or header_row

        for row_idx in range(header_row + 1, max_row + 1):
            brand = str(cell(row_idx, col_brand) or cell(row_idx, col_name) or "").strip()
            article = str(cell(row_idx, col_art) or cell(row_idx, col_name) or brand).strip()
            if _wb_xlsx_skip_row(brand, article):
                continue

            sku_label = _wb_xlsx_sku_label(brand, article)
            sales = _wb_xlsx_safe_float(cell(row_idx, col_sales))
            cost = _wb_xlsx_safe_float(cell(row_idx, col_cost))
            delivery = abs(_wb_xlsx_safe_float(cell(row_idx, col_delivery)))
            returns_qty = int(_wb_xlsx_safe_float(cell(row_idx, col_returns_qty)))
            orders_qty = int(_wb_xlsx_safe_float(cell(row_idx, col_orders_qty)))
            ad_spend = abs(_wb_xlsx_safe_float(cell(row_idx, col_ad_spend)))
            stock = int(_wb_xlsx_safe_float(cell(row_idx, col_stock)))
            daily_sales = _wb_xlsx_safe_float(cell(row_idx, col_daily_sales))

            if sales == 0 and returns_qty == 0 and orders_qty == 0 and stock == 0:
                continue

            bucket = sku_data.setdefault(
                sku_label,
                {
                    "revenue": 0.0,
                    "cost": 0.0,
                    "delivery": 0.0,
                    "returns_count": 0,
                    "orders_count": 0,
                    "ad_spend": 0.0,
                    "stock": 0,
                    "daily_sales": 0.0,
                },
            )
            bucket["revenue"] = float(bucket["revenue"]) + sales
            bucket["cost"] = float(bucket["cost"]) + cost
            bucket["delivery"] = float(bucket["delivery"]) + delivery
            bucket["returns_count"] = int(bucket["returns_count"]) + returns_qty
            bucket["orders_count"] = int(bucket["orders_count"]) + orders_qty
            bucket["ad_spend"] = float(bucket["ad_spend"]) + ad_spend
            bucket["stock"] = max(int(bucket["stock"]), stock)
            bucket["daily_sales"] = max(float(bucket["daily_sales"]), daily_sales)

        if not sku_data:
            return None

        for metrics in sku_data.values():
            orders = int(metrics["orders_count"])
            returns = int(metrics["returns_count"])
            if orders > 0:
                metrics["returns_count"] = min(returns, orders)
            metrics["returns_count"] = int(
                clamp_shop_returns_qty(
                    float(metrics["returns_count"]),
                    sales_qty=float(orders),
                    deliveries_qty=float(orders),
                )
            )

        total_revenue = sum(float(m["revenue"]) for m in sku_data.values())
        if total_revenue <= 0:
            return None

        total_ad_spend = sum(float(m["ad_spend"]) for m in sku_data.values())
        tax_usn = round(total_revenue * _USN_RATE, 2)

        sorted_skus: list[dict[str, Any]] = []
        total_profit = 0.0
        return_logistics_lines: list[str] = []
        total_return_loss = 0.0

        for sku, metrics in sku_data.items():
            orders_count = int(metrics["orders_count"])
            returns_count = int(metrics["returns_count"])
            delivery_rub = float(metrics["delivery"])
            revenue = float(metrics["revenue"])
            cost = float(metrics["cost"])
            ad_spend = float(metrics["ad_spend"])

            total_actions = orders_count + returns_count
            buyout_rate = (
                (orders_count / total_actions * 100.0) if total_actions > 0 else 100.0
            )
            return_loss_rub = _wb_xlsx_return_loss_rub(returns_count, delivery_rub)
            total_return_loss += return_loss_rub
            if returns_count > 0 and return_loss_rub > 0:
                loss_s = f"{return_loss_rub:,.2f}".replace(",", " ")
                return_logistics_lines.append(
                    f"Логистика возвратов: {sku}: {returns_count} возвратов. "
                    f"Общий убыток на пустых покатушках: ≈ {loss_s} руб."
                )

            sku_profit = revenue - cost - delivery_rub - (revenue * _USN_RATE) - ad_spend
            total_profit += sku_profit
            sorted_skus.append(
                {
                    "sku": sku,
                    "revenue": revenue,
                    "profit": sku_profit,
                    "buyout_rate": buyout_rate,
                    "returns_count": returns_count,
                    "return_loss_rub": return_loss_rub,
                    "stock": int(metrics["stock"]),
                    "daily_sales": float(metrics["daily_sales"]),
                }
            )

        sorted_skus.sort(key=lambda x: x["revenue"], reverse=True)

        group_a: list[str] = []
        group_c: list[str] = []
        ballast: list[dict[str, Any]] = []
        non_liquid: list[dict[str, Any]] = []
        oos_predictions: dict[str, int] = {}

        running_sum = 0.0
        for item in sorted_skus:
            running_sum += item["revenue"]
            if running_sum <= total_revenue * 0.8 and item["profit"] > 0:
                group_a.append(item["sku"])
            else:
                group_c.append(item["sku"])

            if item["buyout_rate"] < _BALLAST_BUYOUT_PCT and item["returns_count"] > 0:
                ballast.append(
                    {
                        "sku": item["sku"],
                        "buyout": round(item["buyout_rate"], 1),
                        "returns": item["returns_count"],
                        "loss": round(item["return_loss_rub"], 2),
                    }
                )
            if item["stock"] > 0 and item["revenue"] == 0:
                unit_cost = (
                    float(metrics["cost"]) / max(int(metrics["orders_count"]), 1)
                    if float(metrics["cost"]) > 0
                    else 0.0
                )
                non_liquid.append(
                    {
                        "sku": item["sku"],
                        "stock": item["stock"],
                        "cost": round(unit_cost, 2),
                        "frozen_capital_rub": round(item["stock"] * unit_cost, 2),
                    }
                )
            if item["daily_sales"] > 0:
                days_left = int(item["stock"] / item["daily_sales"])
                if days_left <= 7:
                    oos_predictions[item["sku"]] = days_left

        drr = (total_ad_spend / total_revenue * 100.0) if total_revenue > 0 else 0.0
        business_score = 10.0
        if total_profit <= 0:
            business_score -= 4.0
        if ballast:
            business_score -= 2.0
        if oos_predictions:
            business_score -= 1.5
        if drr > _DRR_WARNING_PCT:
            business_score -= 1.5
        business_score = max(1.0, min(10.0, business_score))

        margin_rate = (total_profit / total_revenue * 100.0) if total_revenue > 0 else 0.0
        shop_avg = (
            total_return_loss / sum(int(m["returns_count"]) for m in sku_data.values())
            if sum(int(m["returns_count"]) for m in sku_data.values()) > 0
            else 0.0
        )
        emoji, status = _business_score_band(business_score)
        verdict = derive_business_verdict(
            business_score=business_score,
            profitability_pct=margin_rate,
            ad_load_pct=drr,
            buyout_coef_pct=0.0,
            worst_unit_label=group_c[0] if group_c else None,
        )
        reason = (
            "📉 Балл занижен из-за балласта и возвратов."
            if ballast
            else "📈 Операционные показатели в пределах нормы."
        )
        traffic_light = {
            "green": (
                "Прибыльные лидеры: " + "; ".join(group_a)
                if group_a
                else "лидеры не выделены"
            ),
            "yellow": (
                f"ДРР {drr:.1f}% — Вы работаете на рекламу, а не на карман."
                if drr > _DRR_WARNING_PCT
                else f"ДРР {drr:.1f}%"
            ),
            "red": (
                "; ".join(
                    f"«{b['sku']}» — убыток ≈ {b['loss']} руб." for b in ballast
                )
                if ballast
                else "критических убытков не выявлено"
            ),
        }

        return {
            "finance": {
                "total_revenue": round(total_revenue, 2),
                "tax_usn": tax_usn,
                "total_profit": round(total_profit, 2),
                "margin_rate": round(margin_rate, 1),
                "drr": round(drr, 1),
                "business_score": round(business_score, 1),
            },
            "health_index": {
                "score": round(business_score, 1),
                "emoji": emoji,
                "status": status,
                "reason": reason,
                "verdict": verdict,
            },
            "abc_analysis": {
                "group_A": group_a
                if group_a
                else ["Лидеры отсутствуют, вся матрица требует санации"],
                "group_C": group_c,
                "total_group_c_count": len(group_c),
            },
            "problem_zones": {
                "ballast": ballast,
                "non_liquid": non_liquid,
            },
            "traffic_light": traffic_light,
            "loss_calculator": {
                "return_logistics": {
                    "total_loss_rub": round(total_return_loss, 2),
                    "lines": [{"text": ln} for ln in return_logistics_lines],
                    "shop_avg_rub_per_unit": round(shop_avg, 2),
                    "min_tariff_rub": MIN_REVERSE_LOGISTICS_RUB,
                },
                "fomo_lost_rub": round(total_return_loss, 2),
            },
            "oos_predictions": oos_predictions,
            "year_forecast_rub": round(total_revenue * 52, 0),
            "localization_index": "не указан в исходных данных",
            "sku_catalog": [
                {
                    "sku": sku,
                    "revenue": round(float(m["revenue"]), 2),
                    "returns_count": int(m["returns_count"]),
                    "orders_count": int(m["orders_count"]),
                }
                for sku, m in sku_data.items()
            ],
            "parser": "wb_weekly_openpyxl_v1",
        }
    finally:
        wb.close()


def build_wb_mpstats_ai_context(
    matrix_rows: list[list[str]],
    *,
    revenue_total: float,
    platform: str | None = "wildberries",
    tax_preset_id: str | None = None,
) -> dict[str, Any]:
    """
    MPSTATS-стиль JSON для OpenRouter: гибридный ETL cfo-v10.

    Все числа считаются в Python (:func:`build_final_metrics_json`);
    LLM только интерпретирует готовый пакет.
    """
    from services.file_processor import build_final_metrics_json, compute_seller_matrix_etl

    if revenue_total <= 0 or not matrix_rows or len(matrix_rows) < 2:
        return {"error": "empty_or_no_revenue", "cfo_build": _FINANCE_REPORT_BUILD}

    final_metrics = build_final_metrics_json(
        matrix_rows,
        revenue_total=revenue_total,
        platform=platform,
        tax_preset_id=tax_preset_id,
    )
    if final_metrics.get("error"):
        return final_metrics

    etl = compute_seller_matrix_etl(matrix_rows, revenue_total=revenue_total, platform=platform)
    wb_metrics = resolve_wb_metrics_for_rows(matrix_rows, revenue_total, platform=platform)
    prompt_metrics = compute_wb_finance_prompt_metrics(
        revenue_total,
        wb_metrics,
        matrix_rows=matrix_rows,
        platform=platform,
        tax_preset_id=tax_preset_id,
    )

    ballast, non_liquid = _extract_problem_zones_structured(etl)
    drr = wb_metrics.ad_load_pct if wb_metrics else float(
        final_metrics.get("shop", {}).get("drr_pct", 0.0)
    )
    business_score = prompt_metrics.business_score if prompt_metrics else 0.0
    verdict = prompt_metrics.verdict if prompt_metrics else ""
    traffic_light = _build_traffic_light_json(
        group_a=list(final_metrics.get("abc_analysis", {}).get("group_A", [])),
        ballast=ballast,
        drr=drr,
        prompt_metrics=prompt_metrics,
        wb_metrics=wb_metrics,
    )
    emoji, status = _business_score_band(business_score)
    reason = (
        _business_score_reason_line(prompt_metrics, wb_metrics)
        if prompt_metrics
        else "Причина оценки рассчитана по операционным метрикам."
    )

    shop = final_metrics.get("shop", {})
    return {
        **final_metrics,
        "finance": {
            "total_revenue": shop.get("total_revenue", revenue_total),
            "tax_usn": shop.get("tax_usn", 0.0),
            "total_profit": shop.get("clear_profit", 0.0),
            "operational_profit": shop.get("operational_profit", 0.0),
            "storage_cost": shop.get("storage_cost", 0.0),
            "credit_deductions": shop.get("credit_deductions", 0.0),
            "margin_rate": shop.get("margin_rate_pct", 0.0),
            "drr": shop.get("drr_pct", drr),
            "business_score": round(business_score, 1),
            "buyout_coef_pct": shop.get("buyout_coef_pct", 0.0),
        },
        "health_index": {
            "score": round(business_score, 1),
            "emoji": emoji,
            "status": status,
            "reason": reason,
            "verdict": verdict,
        },
        "problem_zones": {
            "ballast": ballast,
            "non_liquid": non_liquid,
        },
        "traffic_light": traffic_light,
        "loss_calculator": {
            "return_logistics": build_mpstats_return_logistics_payload(
                etl, revenue_total=revenue_total
            ),
            "fomo_lost_rub": round(prompt_metrics.fomo_lost_rub, 2) if prompt_metrics else 0.0,
        },
        "localization_index": (
            prompt_metrics.localization_index_line if prompt_metrics else "не указан в исходных данных"
        ),
        "strategic_plan_hints": _build_strategic_plan_lines(prompt_metrics, wb_metrics)
        if prompt_metrics
        else [],
    }


def resolve_wb_revenue_total(
    *,
    calculated_total: float,
    file_path: str | Path | None = None,
    matrix_rows: list[list[str]] | None = None,
    platform: str | None = "wildberries",
) -> float:
    """Выручка для CFO: worker → openpyxl JSON → preprocess матрицы."""
    if calculated_total > 0:
        return float(calculated_total)
    if file_path is not None:
        try:
            raw = prepare_wb_data_for_ai(file_path, platform=platform or "wildberries")
            data = json.loads(raw)
            rev = float((data.get("finance") or {}).get("total_revenue") or 0.0)
            if rev > 0:
                return rev
        except Exception:
            logger.debug("resolve_wb_revenue_total: file parse failed", exc_info=True)
    if matrix_rows:
        from services.table_xlsx_preprocess import preprocess_xlsx_rows

        pre = preprocess_xlsx_rows(matrix_rows, title="WB")
        if pre.revenue_total > 0:
            return float(pre.revenue_total)
    return 0.0


def _is_publishable_wb_finance_html(html: str) -> bool:
    """Ответ OpenRouter пригоден к показу (не пустой и похож на CFO-отчёт)."""
    sample = (html or "").strip()
    if len(sample) < 80:
        return False
    low = sample.lower()
    return (
        "финансовый экспресс" in low
        or "abc-анализ" in low
        or "индекс здоровья" in low
    )


def _build_local_wb_finance_html(
    revenue_total: float,
    wb_metrics: WbMarketplaceMetrics | None,
    *,
    matrix_rows: list[list[str]] | None = None,
    platform: str | None = None,
    tax_preset_id: str | None = None,
) -> str | None:
    """Гарантированный локальный CFO-отчёт (без OpenRouter)."""
    if revenue_total <= 0:
        return None
    if matrix_rows and len(matrix_rows) >= 2:
        from services.audit_tax import resolve_audit_tax_preset
        from services.file_processor import build_cfo_metrics_dict_from_rows

        preset = resolve_audit_tax_preset(tax_preset_id)
        cfo_metrics = build_cfo_metrics_dict_from_rows(
            matrix_rows,
            platform or "wildberries",
            preset.regime,
            preset.rate_percent,
        )
        if not cfo_metrics.get("error"):
            return append_wb_finance_mini_app_cta(
                build_wb_finance_consulting_html_from_cfo_metrics(cfo_metrics)
            )
    prompt_metrics = compute_wb_finance_prompt_metrics(
        revenue_total,
        wb_metrics,
        matrix_rows=matrix_rows,
        platform=platform,
        tax_preset_id=tax_preset_id,
    )
    if prompt_metrics is None:
        return None
    return append_wb_finance_mini_app_cta(
        build_wb_finance_express_html_local(prompt_metrics, wb_metrics)
    )


def resolve_wb_mpstats_context(
    *,
    file_path: str | Path | None = None,
    matrix_rows: list[list[str]] | None = None,
    revenue_total: float = 0.0,
    platform: str | None = "wildberries",
    title: str | None = None,
) -> dict[str, Any]:
    """Единая точка: Excel или матрица → MPSTATS JSON-словарь (все расчёты в Python)."""
    if file_path is not None:
        try:
            raw = prepare_wb_data_for_ai(file_path, platform=platform or "wildberries", title=title)
            data = json.loads(raw)
            if isinstance(data, dict) and not data.get("error"):
                return data
        except Exception:
            logger.exception("resolve_wb_mpstats_context: prepare_wb_data_for_ai failed")
    rev = float(revenue_total or 0.0)
    if matrix_rows and len(matrix_rows) >= 2:
        if rev <= 0:
            from services.table_xlsx_preprocess import preprocess_xlsx_rows

            pre = preprocess_xlsx_rows(matrix_rows, title=title or "WB")
            rev = float(pre.revenue_total or 0.0)
        if rev > 0:
            return build_wb_mpstats_ai_context(
                matrix_rows,
                revenue_total=rev,
                platform=platform,
            )
    return {"error": "empty_or_no_revenue"}


def build_wb_finance_json_user_message(json_payload: str | dict[str, Any]) -> str:
    """User-сообщение OpenRouter: только готовый final_metrics_json (cfo-v10)."""
    if isinstance(json_payload, dict):
        body = json.dumps(json_payload, ensure_ascii=False, indent=2)
    else:
        body = json_payload.strip()
    return (
        "Ты — CFO. На основе готовых выверенных математических данных из JSON "
        "сформируй бизнес-выводы, Светофор эффективности и План действий. "
        "Не пересчитывай цифры, бери их строго из JSON. "
        "Строго упакуй JSON в HTML-отчёт по структуре из system prompt. "
        "Запрещено менять числа, дополнять товары, сокращать списки и использовать слово «ИИ».\n\n"
        f"{body}"
    )


def prepare_wb_data_for_ai(
    file_path: str | Path,
    *,
    platform: str = "wildberries",
    title: str | None = None,
) -> str:
    """
    Считывает Excel Wildberries, считает метрики в стиле MPSTATS и возвращает JSON для OpenRouter.

    Сначала — парсер еженедельного отчёта WB (openpyxl, агрегация по SKU),
    при неудаче — матричный ETL из preprocess_xlsx_rows.
    """
    from pathlib import Path as PathCls

    from services.file_processor import read_xlsx_rows_from_path
    from services.table_xlsx_preprocess import preprocess_xlsx_rows

    path = PathCls(file_path)

    if platform in (None, "", "wildberries"):
        weekly = build_wb_weekly_xlsx_ai_context(path)
        if weekly is not None:
            return json.dumps(weekly, ensure_ascii=False, indent=2)

    raw_rows = read_xlsx_rows_from_path(path)
    display_title = title or path.stem or "Отчёт WB"
    pre = preprocess_xlsx_rows(raw_rows, title=display_title)
    context = build_wb_mpstats_ai_context(
        pre.rows,
        revenue_total=float(pre.revenue_total or 0.0),
        platform=platform,
    )
    return json.dumps(context, ensure_ascii=False, indent=2)


def _format_sku_bullet_lines(
    buyout_pct: float,
    revenue: float,
    *,
    stock_qty: float,
    sales_qty: float,
    days_until_stockout: float | None,
) -> str:
    """«ballast» — низкий выкуп; «illiquid» — залежалый остаток без спроса."""
    if buyout_pct < _BALLAST_BUYOUT_PCT and sales_qty > 0:
        return "ballast"
    if sales_qty <= 0 and stock_qty >= _ILLIQUID_MIN_STOCK:
        return "illiquid"
    if (
        revenue <= 0
        and stock_qty >= _ILLIQUID_MIN_STOCK
        and buyout_pct <= _CRITICAL_BUYOUT_PCT
    ):
        return "illiquid"
    if (
        days_until_stockout is not None
        and days_until_stockout > _SLOW_TURNOVER_DAYS
        and sales_qty > 0
        and stock_qty >= _ILLIQUID_MIN_STOCK
    ):
        return "illiquid"
    if buyout_pct < _BALLAST_BUYOUT_PCT:
        return "ballast"
    if revenue <= 0 and buyout_pct < 55.0:
        return "ballast"
    return "ballast" if buyout_pct < 55.0 else "illiquid"


def _ballast_reason(buyout_pct: float) -> str:
    return (
        f"выкуп {buyout_pct:.1f}% (менее {_BALLAST_BUYOUT_PCT:.0f}%) — "
        "покатушки и пустая обратная логистика съедают маржу"
    )


def _illiquid_reason(
    *,
    stock_qty: float,
    sales_qty: float,
    days_until_stockout: float | None,
    revenue: float,
) -> str:
    if sales_qty <= 0 and stock_qty > 0:
        stock_s = f"{stock_qty:.0f}".replace(",", " ")
        return f"нет продаж за период, на складе {stock_s} шт. — капитал заморожен"
    if days_until_stockout is not None and days_until_stockout > _SLOW_TURNOVER_DAYS:
        return (
            f"медленная оборачиваемость (~{days_until_stockout:.0f} дн. запаса), "
            f"выручка {_fmt_rub_in_code(revenue)} руб."
        )
    stock_s = f"{stock_qty:.0f}".replace(",", " ")
    return f"залежалый остаток ({stock_s} шт.) без достаточного спроса"


def build_matrix_problem_zones_block(matrix_etl: object | None) -> str:
    """Текст подсказки ETL: балласт и неликвид из группы C (с агрегацией хвоста)."""
    return build_matrix_problem_zones_block_from_aggregation(
        aggregate_matrix_display(matrix_etl)
    )


def _format_sku_bullet_lines(
    items: Iterable[tuple[str, str]],
    *,
    max_items: int | None = None,
    overflow_suffix: str | None = None,
) -> str:
    """Список товаров маркерами «•», по одному на строку (без сокращений по умолчанию)."""
    lines: list[str] = []
    batch = [(n, a) for n, a in items if _sku_label_is_valid(n, a)]
    display = batch if max_items is None else batch[:max_items]
    for name, article in display:
        lines.append(f"• {_format_sku_label(name, article)}")
    if overflow_suffix:
        lines.append(overflow_suffix)
    elif max_items is not None and len(batch) > max_items:
        lines.append(_ru_more_goods_suffix(len(batch) - max_items))
    return "\n".join(lines) if lines else "• убыточных товаров не выявлено"


def _expand_fomo_breakdown(parts: tuple[str, ...]) -> tuple[str, ...]:
    """Разбивает строки с «;» на отдельные пункты для маркированного списка."""
    expanded: list[str] = []
    for part in parts:
        chunk = (part or "").strip()
        if not chunk:
            continue
        if chunk.startswith("Логистика невыкупленных:"):
            chunk = chunk.removeprefix("Логистика невыкупленных:").strip()
        if chunk.startswith("Логистика возвратов:"):
            expanded.append(chunk)
            continue
        if "; " in chunk:
            expanded.extend(p.strip() for p in chunk.split("; ") if p.strip())
        else:
            expanded.append(chunk)
    return tuple(expanded)


def _leader_buyout_is_healthy(buyout_pct: float, margin_rub: float = 1.0) -> bool:
    return _sku_qualifies_for_success_zone(buyout_pct, margin_rub)


def _format_fomo_details_block(parts: tuple[str, ...]) -> str:
    expanded = _expand_fomo_breakdown(parts)
    if not expanded:
        return "• критических зон упущенной выгоды не выявлено"
    return "\n".join(f"• {p}" for p in expanded[:6])


def _build_strategic_plan_lines(
    prompt_metrics: WbFinancePromptMetrics,
    wb_metrics: WbMarketplaceMetrics | None,
) -> list[str]:
    """Три шага плана: масштабирование, убыточный SKU, OOS — только из ETL."""
    lines: list[str] = []

    scale_item: dict[str, Any] | None = None
    for item in prompt_metrics.abc_group_a_items:
        if float(item.get("unit_profit_rub", 0.0)) > 0:
            scale_item = item
            break
    if scale_item is None:
        for item in sorted(
            prompt_metrics.sku_catalog_items,
            key=lambda x: float(x.get("margin_rub", 0.0)),
            reverse=True,
        ):
            if float(item.get("margin_rub", 0.0)) > 0:
                scale_item = {
                    "label": _format_sku_label(
                        str(item.get("name", "")),
                        str(item.get("article_id", "")),
                    ),
                    "unit_profit_rub": float(item.get("margin_rub", 0.0)),
                }
                break

    if scale_item:
        scale_label = _escape_verdict(str(scale_item.get("label", "—")))
        lines.append(
            f"<b>1.</b> Усилить закуп и рекламу на <b>{scale_label}</b> — "
            f"чистая прибыль с одной продажи: "
            f"<code>{_fmt_rub_in_code(float(scale_item.get('unit_profit_rub', 0.0)))}</code> руб."
        )
    else:
        lines.append(
            "<b>1.</b> Контролируйте ДРР и окупаемость рекламы — "
            "прибыльный драйвер для масштабирования в этом периоде не выделен."
        )

    if prompt_metrics.loss_sku_items:
        worst = prompt_metrics.loss_sku_items[0]
        worst_label = _escape_verdict(str(worst.get("label", "—")))
        lines.append(
            f"<b>2.</b> Для <b>{worst_label}</b> поднимите цену "
            f"или остановите рекламу — убыток "
            f"<code>{_fmt_rub_in_code(abs(float(worst.get('net_profit_rub', 0.0))))}</code> руб."
        )
    elif prompt_metrics.adv_load_pct >= 30:
        lines.append(
            f"<b>2.</b> Катастрофический ДРР <code>{prompt_metrics.adv_load_pct:.1f}%</code> — "
            "немедленно режьте рекламный бюджет минимум вдвое, отключите все неокупаемые кампании "
            "и пересоберите семантику под маржинальные товары."
        )
    elif prompt_metrics.adv_load_pct > _HIGH_DRR_PCT:
        lines.append(
            f"<b>2.</b> Срочно <b>снизить ДРР</b> с <code>{prompt_metrics.adv_load_pct:.1f}%</code> "
            "до 12–15%: отключите неокупаемые кампании, сузьте ключевые слова и ставки."
        )
    else:
        lines.append(
            f"<b>2.</b> Держите ДРР не выше <code>15%</code> — еженедельно чистите неокупаемые ключи."
        )

    if prompt_metrics.oos_zero_stock_items:
        oos_names = ", ".join(
            _escape_verdict(_oos_item_display_label(dict(item)))
            for item in prompt_metrics.oos_zero_stock_items[:3]
        )
        lines.append(
            f"<b>3.</b> Срочно закупите: <b>{oos_names}</b> — "
            "товар закончился на складе при сохранённом спросе."
        )
    elif prompt_metrics.oos_critical_stock_items:
        crit_names = ", ".join(
            _escape_verdict(_oos_item_display_label(dict(item)))
            for item in prompt_metrics.oos_critical_stock_items[:3]
        )
        lines.append(
            f"<b>3.</b> Запланируйте поставку в течение 48 часов: <b>{crit_names}</b> — "
            "остаток на складе критически мал при текущем темпе продаж."
        )
    elif prompt_metrics.oos_forecast_summary:
        lines.append(
            f"<b>3.</b> Контролируйте остатки по рисковым артикулам: "
            f"<i>{_escape_verdict(prompt_metrics.oos_forecast_summary)}</i>"
        )
    else:
        lines.append(
            "<b>3.</b> Еженедельно сверяйте остатки FBO/FBS и не допускайте обнуления "
            "по товарам с продажами в отчётном периоде."
        )
    return lines
_WB_FINANCE_MAX_OUTPUT_TOKENS = 1400
_WB_FINANCE_TELEGRAM_SOFT_MAX_CHARS = 2000
_FINANCE_SEPARATOR = "────────────────────────"
# Подпись сборки — единый источник: services.table_text_response.FINANCE_REPORT_BUILD
_OLD_FINALE_MARKERS = (
    "Финальный Excel",
    "интерактивный дашборд",
    "Автопилот по API",
    "Хватит загружать",
)
_LEGACY_FINANCE_MARKERS = (
    "ИИ-ПЛАН",
    "ИИ-Инсайт",
    "ИИ-ИНСАЙТ",
    "Senior ИИ",
    "Серверный расчёт",
    "Серверный",
    "локальный ETL",
    "ИИ-Моделирование",
)
_LEGACY_FINANCE_REPLACEMENTS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"ИИ[\s\-–—]*План\s*действий", re.IGNORECASE), "ПЛАН ДЕЙСТВИЙ ДЛЯ ПРЕДПРИНИМАТЕЛЯ НА СЕГОДНЯ"),
    (re.compile(r"ИИ[\s\-–—]*Инсайт", re.IGNORECASE), "ГЛАВНЫЙ АНАЛИТИЧЕСКИЙ ВЫВОД"),
    (re.compile(r"КЛЮЧЕВОЙ\s+БИЗНЕС-ВЕРДИКТ", re.IGNORECASE), "ГЛАВНЫЙ АНАЛИТИЧЕСКИЙ ВЫВОД"),
    (re.compile(r"ГЛАВНЫЙ\s+ВЫВОД\s+ИИ", re.IGNORECASE), "ГЛАВНЫЙ АНАЛИТИЧЕСКИЙ ВЫВОД"),
    (re.compile(r"ВАЛОВАЯ\s+ВЫРУЧКА", re.IGNORECASE), "ОБЩАЯ ВЫРУЧКА"),
    (re.compile(r"Серверный\s+расч[её]т", re.IGNORECASE), "Потенциально можно вернуть в оборот"),
    (re.compile(r"ABC[\s\-–—]*АНАЛИЗ\s+МАТРИЦЫ\s*\(\s*локальный\s+ETL[^)]*\)", re.IGNORECASE), "ABC-АНАЛИЗ ПРОДАЖ"),
    (re.compile(r"ABC[\s\-–—]*АНАЛИЗ\s+МАТРИЦЫ", re.IGNORECASE), "ABC-АНАЛИЗ ПРОДАЖ"),
    (re.compile(r"РЕЙТИНГ\s+ПРОДАЖ\s+ПО\s+ТОВАРАМ", re.IGNORECASE), "ABC-АНАЛИЗ ПРОДАЖ"),
    (re.compile(r"БИЗНЕС-СКОРИНГ\s+МАГАЗИНА", re.IGNORECASE), "ИНДЕКС ЗДОРОВЬЯ БИЗНЕСА"),
    (re.compile(r"ИНДЕКС\s+ЗДОРОВЬЯ\s+МАГАЗИНА", re.IGNORECASE), "ИНДЕКС ЗДОРОВЬЯ БИЗНЕСА"),
    (re.compile(r"СВЕТОФОР\s+ЗДОРОВЬЯ\s+БИЗНЕСА", re.IGNORECASE), "СВЕТОФОР ЭФФЕКТИВНОСТИ"),
    (re.compile(r"Проблемные\s+зоны\s+матрицы", re.IGNORECASE), "Проблемные зоны и скрытые убытки"),
    (re.compile(r"ПРОГНОЗ\s+И\s+КРЭШ-ТЕСТ", re.IGNORECASE), "ПРОГНОЗ И ОБНУЛЕНИЕ ОСТАТКОВ"),
    (re.compile(r"СТРАТЕГИЧЕСКИЙ\s+ПЛАН\s+ДЕЙСТВИЙ", re.IGNORECASE), "ПЛАН ДЕЙСТВИЙ ДЛЯ ПРЕДПРИНИМАТЕЛЯ"),
    (re.compile(r"\bOOS\b", re.IGNORECASE), "Обнуление остатков на складе"),
    (re.compile(r"риск\s+OOS", re.IGNORECASE), "риск обнуления остатков"),
    (
        re.compile(r"\(\s*0\s*шт\.?\s*—\s*ЗАКОНЧИЛСЯ\s*\)", re.IGNORECASE),
        "— 🔴 ТОВАР ПОЛНОСТЬЮ ЗАКОНЧИЛСЯ",
    ),
    (
        re.compile(
            r"остаток\s+\d+\s*шт\.?\s*\(\s*ЗАКОНЧИТСЯ\s+через\s+\d+\s*дн\.?\s*\)",
            re.IGNORECASE,
        ),
        "— 🟡 СКОРО ЗАКОНЧИТСЯ (критический уровень запасов)",
    ),
    (re.compile(r"cfo-v11\.2", re.IGNORECASE), _FINANCE_REPORT_BUILD),
)
_TECH_HINT_PAREN_RE = re.compile(
    r"\(\s*[^)]*(?:"
    r"локальный\s+ETL|не\s+пересчитывай|"
    r"по\s+2[–\-]3\s+предложен|"
    r"каждый\s+шаг|каждый\s+SKU|каждый\s+источник|"
    r"маркер\s+«•»|"
    r"без\s+воды|"
    r"строго\s+из\s+шаблона"
    r")[^)]*\)",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class WbFinancePromptMetrics:
    """Числовые метрики для подстановки в системный промпт."""

    revenue: float
    tax: float
    clear_profit: float
    adv_load_pct: float
    buy_ratio_pct: float
    year_forecast: float
    profitability_pct: float
    business_score: float
    verdict: str
    fomo_lost_rub: float
    fomo_breakdown: tuple[str, ...]
    logistics_fomo_rub: float = 0.0
    abc_a_leader: str = "—"
    abc_a_leader_name: str = "—"
    abc_a_leader_article: str = "—"
    abc_a_leader_revenue: float = 0.0
    abc_a_leader_margin: float = 0.0
    abc_a_leader_buyout: float = 0.0
    abc_a_count: int = 0
    abc_c_count: int = 0
    abc_c_summary: str = "нет"
    outsider_name: str = "—"
    outsider_article: str = "—"
    outsider_loss: float = 0.0
    outsider_buyout: float = 0.0
    sku_catalog_lines: tuple[str, ...] = ()
    sku_catalog_items: tuple[dict[str, Any], ...] = ()
    oos_forecast_line: str = "данных по остаткам недостаточно"
    oos_forecast_summary: str = ""
    total_ad_cost: float = 0.0
    storage_cost: float = 0.0
    credit_deductions: float = 0.0
    operational_profit: float = 0.0
    sales_qty: float = 0.0
    returns_qty: float = 0.0
    deliveries_qty: float = 0.0
    reverse_logistics_shop_avg: float = 0.0
    return_logistics_block: str = "• существенных потерь на обратной логистике не выявлено"
    matrix_problem_zones_block: str = "проблемных зон в группе C не выявлено"
    localization_index_line: str = "не указан в исходных данных"
    abc_group_a_items: tuple[dict[str, Any], ...] = ()
    loss_sku_items: tuple[dict[str, Any], ...] = ()
    non_liquid_items: tuple[dict[str, Any], ...] = ()
    oos_zero_stock_items: tuple[dict[str, Any], ...] = ()
    oos_critical_stock_items: tuple[dict[str, Any], ...] = ()
    non_liquid_frozen_total_rub: float = 0.0
    matrix_aggregation: WbFinanceMatrixAggregation = field(
        default_factory=WbFinanceMatrixAggregation
    )


def compute_business_score(
    *,
    profitability_pct: float,
    ad_load_pct: float,
    buyout_coef_pct: float,
    worst_unit_net: float | None,
) -> float:
    """Бизнес-скоринг 1.0–10.0 на основе рентабельности и операционных рисков."""
    score = 5.0
    if profitability_pct >= 18:
        score += 2.5
    elif profitability_pct >= 12:
        score += 1.5
    elif profitability_pct >= 7:
        score += 0.5
    elif profitability_pct < 4:
        score -= 2.0
    elif profitability_pct < 7:
        score -= 0.5

    if buyout_coef_pct >= 72:
        score += 1.0
    elif buyout_coef_pct >= 58:
        score += 0.3
    elif 0 < buyout_coef_pct < 45:
        score -= 1.5
    elif 0 < buyout_coef_pct < 55:
        score -= 0.7

    if ad_load_pct > 28:
        score -= 1.5
    elif ad_load_pct > 20:
        score -= 0.8
    elif 0 < ad_load_pct <= 12:
        score += 0.4

    if worst_unit_net is not None and worst_unit_net < 0:
        score -= 1.2

    return round(max(1.0, min(10.0, score)), 1)


def _business_score_band(score: float) -> tuple[str, str]:
    """Эмодзи и человекочитаемый статус по шкале 1–10."""
    if score < 5.0:
        return (
            "🔴",
            "КРИТИЧЕСКИЙ УРОВЕНЬ — Высокий риск кассового разрыва",
        )
    if score < 8.0:
        return (
            "🟡",
            "НОРМАЛЬНЫЙ УРОВЕНЬ — Требуется оптимизация расходов",
        )
    return "🟢", "ОТЛИЧНЫЙ УРОВЕНЬ — Эффективное управление"


def _business_score_reason_line(
    prompt_metrics: WbFinancePromptMetrics,
    wb_metrics: WbMarketplaceMetrics | None,
) -> str:
    """Одна строка причины занижения/завышения балла."""
    reasons_low: list[str] = []
    reasons_high: list[str] = []

    if prompt_metrics.profitability_pct < 7:
        reasons_low.append("низкой рентабельности")
    elif prompt_metrics.profitability_pct >= 18:
        reasons_high.append("высокой маржинальности")

    if 0 < prompt_metrics.buy_ratio_pct < 45:
        reasons_low.append("критического выкупа")
    elif prompt_metrics.buy_ratio_pct >= 72:
        reasons_high.append("стабильного выкупа")

    if prompt_metrics.adv_load_pct > 20:
        reasons_low.append("раздутого ДРР")
    elif 0 < prompt_metrics.adv_load_pct <= 12:
        reasons_high.append("контролируемой рекламы")

    if prompt_metrics.outsider_loss > 0:
        reasons_low.append("убыточных SKU")

    if wb_metrics and wb_metrics.top5_units:
        worst = min(wb_metrics.top5_units, key=lambda u: u.net_income)
        if worst.net_income < 0:
            reasons_low.append("убыточных позиций в матрице")

    if prompt_metrics.business_score >= 7.0 and reasons_high:
        return f"📈 Балл высокий благодаря {', '.join(reasons_high)}."
    if reasons_low:
        return f"📉 Балл занижен из-за {', '.join(reasons_low)}."
    if prompt_metrics.business_score >= 8.0:
        return "📈 Балл высокий благодаря сбалансированной экономике магазина."
    return "📉 Балл занижен из-за операционных рисков — нужна оптимизация."


def derive_business_verdict(
    *,
    business_score: float,
    profitability_pct: float,
    ad_load_pct: float,
    buyout_coef_pct: float,
    worst_unit_label: str | None,
) -> str:
    """Ёмкий вердикт для блока бизнес-скоринга."""
    if business_score >= 8.0:
        return "Высокая маржинальность — фокус на масштабировании лидеров ассортимента."
    if business_score >= 6.5:
        if ad_load_pct > 22:
            return "Стабильная база при риске перерасхода на рекламу и просадки кассы."
        return "Здоровый бизнес с точками роста в операционной эффективности."
    if profitability_pct >= 10 and buyout_coef_pct > 0 and buyout_coef_pct < 52:
        return "Высокая маржинальность при критическом риске кассового разрыва из-за низкого выкупа."
    if worst_unit_label and ad_load_pct > 20:
        return (
            f"Критический риск кассового разрыва: «{worst_unit_label[:32]}» и реклама "
            f"вымывают оборотные средства."
        )
    if profitability_pct < 5:
        return "Критически низкая рентабельность — угроза кассового разрыва в ближайшем цикле."
    return "Требуется жёсткая оптимизация убыточных SKU, логистики и рекламного ДРР."


def compute_fomo_lost_rub(
    revenue_total: float,
    wb_metrics: WbMarketplaceMetrics | None,
) -> tuple[float, tuple[str, ...]]:
    """
    Калькулятор упущенной выгоды (руб.): низкий выкуп, реклама, убыточные SKU.

    Возвращает ``(сумма, расшифровка для промпта)``.
    """
    if revenue_total <= 0 or wb_metrics is None:
        return 0.0, ()

    parts: list[str] = []
    total = 0.0

    buyout = wb_metrics.buyout_coef_pct
    if buyout > 0 and buyout < 65:
        gap = 65.0 - buyout
        buyout_loss = revenue_total * (gap / 100.0) * 0.22
        if buyout_loss > 50:
            total += buyout_loss
            parts.append(
                f"низкий выкуп {buyout:.1f}% → логистика возвратов и покатушки ≈ "
                f"{_fmt_rub_in_code(buyout_loss)} руб."
            )

    if wb_metrics.returns_qty > 0 and wb_metrics.deliveries_qty > 0:
        validated_returns = clamp_shop_returns_qty(
            wb_metrics.returns_qty,
            sales_qty=wb_metrics.sales_qty,
            deliveries_qty=wb_metrics.deliveries_qty,
        )
        ret_rate = validated_returns / wb_metrics.deliveries_qty
        if ret_rate > 0.12 and validated_returns > 0:
            unit = max(MIN_REVERSE_LOGISTICS_RUB, revenue_total * ret_rate * 0.08 / validated_returns)
            penalty = validated_returns * unit
            if penalty > 30:
                total += penalty
                parts.append(
                    f"возвраты {validated_returns:.0f} шт. → обратная логистика ≈ "
                    f"{_fmt_rub_in_code(penalty)} руб."
                )

    if wb_metrics.ad_load_pct > 18 and wb_metrics.total_advertising_cost > 0:
        waste = wb_metrics.total_advertising_cost * min(0.45, (wb_metrics.ad_load_pct - 15) / 100.0)
        if waste > 50:
            total += waste
            parts.append(
                f"рекламный перерасход ДРР {wb_metrics.ad_load_pct:.1f}% ≈ "
                f"{_fmt_rub_in_code(waste)} руб."
            )

    for unit in wb_metrics.top5_units:
        if unit.net_income < 0:
            sku_loss = abs(unit.net_income) * max(8.0, wb_metrics.sales_qty / max(len(wb_metrics.top5_units), 1))
            if sku_loss > 20:
                total += sku_loss
                parts.append(
                    f"убыточный SKU «{unit.label[:28]}» "
                    f"(<code>{_fmt_rub_in_code(unit.net_income)}</code>/шт.) ≈ "
                    f"{_fmt_rub_in_code(sku_loss)} руб."
                )
                break

    return round(total, 2), tuple(parts)


def compute_wb_finance_prompt_metrics(
    revenue_total: float,
    wb_metrics: WbMarketplaceMetrics | None = None,
    *,
    matrix_rows: list[list[str]] | None = None,
    platform: str | None = None,
    tax_preset_id: str | None = None,
) -> WbFinancePromptMetrics | None:
    """Собирает переменные ETL для system/user prompt и локального fallback."""
    if revenue_total <= 0:
        return None

    from services.file_processor import aggregate_cfo_engine_v11_1, compute_seller_matrix_etl

    matrix_etl = (
        compute_seller_matrix_etl(
            matrix_rows,
            revenue_total=revenue_total,
            platform=platform,
        )
        if matrix_rows
        else None
    )
    engine = (
        aggregate_cfo_engine_v11_1(
            matrix_rows, platform=platform, tax_preset_id=tax_preset_id
        )
        if matrix_rows
        else None
    )
    storage_cost = wb_metrics.storage_cost if wb_metrics else 0.0
    credit_deductions = wb_metrics.credit_deductions if wb_metrics else 0.0
    ad_cost = wb_metrics.total_advertising_cost if wb_metrics else 0.0
    logistics_cost = wb_metrics.logistics_cost if wb_metrics else 0.0
    commission_cost = wb_metrics.commission_cost if wb_metrics else 0.0
    other_deductions = wb_metrics.other_deductions if wb_metrics else 0.0
    if engine is not None:
        storage_cost = engine.total_storage_cost
        credit_deductions = engine.credit_deductions
        ad_cost = engine.total_ad_spend
        logistics_cost = engine.logistics_cost
        commission_cost = engine.commission_cost
        other_deductions = engine.total_system_losses
    cost_of_goods = 0.0
    if matrix_etl:
        for detail in matrix_etl.sku_catalog:
            if detail.sales_qty > 0 and detail.unit_cost_rub > 0:
                cost_of_goods += detail.unit_cost_rub * detail.sales_qty
            elif detail.unit_cost_rub > 0:
                cost_of_goods += detail.unit_cost_rub
    tax_base = (
        engine.tax_base_revenue
        if engine is not None and engine.tax_base_revenue > 0
        else revenue_total
    )
    tax = engine.tax_total if engine is not None else tax_base * _USN_RATE
    if engine is not None:
        operational_profit = engine.operational_profit
        clear_profit = engine.clear_profit
    else:
        operational_profit = (
            revenue_total
            - cost_of_goods
            - storage_cost
            - ad_cost
            - logistics_cost
            - commission_cost
            - other_deductions
            - tax
        )
        clear_profit = operational_profit - credit_deductions
    profitability = (clear_profit / tax_base * 100.0) if tax_base > 0 else 0.0
    adv_load = wb_metrics.ad_load_pct if wb_metrics else 0.0
    buy_ratio = wb_metrics.buyout_coef_pct if wb_metrics else 0.0
    worst_unit = None
    worst_label = None
    if wb_metrics and wb_metrics.top5_units:
        worst_unit = min(wb_metrics.top5_units, key=lambda u: u.net_income)
        if worst_unit.net_income < 0:
            worst_label = worst_unit.label
    business_score = compute_business_score(
        profitability_pct=profitability,
        ad_load_pct=adv_load,
        buyout_coef_pct=buy_ratio,
        worst_unit_net=worst_unit.net_income if worst_unit else None,
    )
    verdict = derive_business_verdict(
        business_score=business_score,
        profitability_pct=profitability,
        ad_load_pct=adv_load,
        buyout_coef_pct=buy_ratio,
        worst_unit_label=worst_label,
    )
    if clear_profit < 0 and operational_profit > 0 and credit_deductions > 0:
        verdict = (
            "Убыток вызван удержанием по кредиту. "
            "Операционная прибыль от продаж при этом положительна."
        )
    fomo_rub, fomo_parts = compute_fomo_lost_rub(revenue_total, wb_metrics)
    logistics_fomo = 0.0
    abc_a_leader = "—"
    abc_a_leader_name = "—"
    abc_a_leader_article = "—"
    abc_a_leader_revenue = 0.0
    abc_a_leader_margin = 0.0
    abc_a_leader_buyout = 0.0
    abc_a_count = 0
    abc_c_count = 0
    abc_c_summary = "нет"
    matrix_problem_zones_block = "проблемных зон в группе C не выявлено"
    localization_index_line = "не указан в исходных данных"
    outsider_name = "—"
    outsider_article = "—"
    outsider_loss = 0.0
    outsider_buyout = 0.0
    sku_catalog_lines: tuple[str, ...] = ()
    sku_catalog_items: tuple[dict[str, Any], ...] = ()
    oos_line = _build_oos_forecast_line(None, (), ())
    oos_summary = ""
    reverse_logistics_shop_avg = 0.0
    return_logistics_block = "• существенных потерь на обратной логистике не выявлено"
    abc_group_a_items: tuple[dict[str, Any], ...] = ()
    loss_sku_items: tuple[dict[str, Any], ...] = ()
    non_liquid_items: tuple[dict[str, Any], ...] = ()
    oos_zero_stock_items: tuple[dict[str, Any], ...] = ()
    oos_critical_stock_items: tuple[dict[str, Any], ...] = ()
    non_liquid_frozen_total_rub = 0.0
    matrix_aggregation = WbFinanceMatrixAggregation()
    if matrix_etl:
        logistics_fomo = matrix_etl.logistics_fomo_rub
        reverse_logistics_shop_avg = matrix_etl.reverse_logistics_shop_avg
        return_logistics_block = matrix_etl.return_logistics_block
        if matrix_etl.logistics_fomo_rub > 0:
            fomo_rub = round(fomo_rub + logistics_fomo, 2)
            if matrix_etl.logistics_fomo_items:
                fomo_parts = _expand_fomo_breakdown(
                    (*fomo_parts, *matrix_etl.logistics_fomo_items)
                )
            else:
                fomo_parts = _expand_fomo_breakdown(
                    (*fomo_parts, matrix_etl.logistics_fomo_detail)
                )
        if matrix_etl.abc_group_a:
            leader = matrix_etl.abc_group_a[0]
            abc_a_leader = leader.name
            abc_a_leader_name = leader.name
            abc_a_leader_article = leader.article_id
            abc_a_leader_revenue = leader.revenue
            abc_a_leader_margin = leader.net_profit
            abc_a_leader_buyout = leader.buyout_pct
        else:
            abc_a_leader = matrix_etl.abc_a_leader
            abc_a_leader_name = matrix_etl.abc_a_leader
            abc_a_leader_article = matrix_etl.abc_a_leader
        abc_a_count = len(matrix_etl.abc_group_a)
        abc_c_count = len(matrix_etl.abc_group_c)
        matrix_aggregation = aggregate_matrix_display(matrix_etl)
        if matrix_aggregation.abc_c_display_lines:
            abc_c_lines = list(matrix_aggregation.abc_c_display_lines)
            if matrix_aggregation.tail_c_count > 0:
                abc_c_lines.append(
                    _tail_line_group_c(
                        matrix_aggregation.tail_c_count,
                        matrix_aggregation.tail_c_revenue,
                    )
                )
            abc_c_summary = "\n".join(abc_c_lines)
        elif abc_c_count == 0:
            abc_c_summary = "убыточных товаров не выявлено"
        else:
            abc_c_summary = "• убыточных товаров не выявлено"
        matrix_problem_zones_block = build_matrix_problem_zones_block_from_aggregation(
            matrix_aggregation
        )
        if matrix_etl.outsider_sku:
            outsider_name = matrix_etl.outsider_sku.name
            outsider_article = matrix_etl.outsider_sku.article_id
            outsider_loss = abs(matrix_etl.outsider_sku.net_profit)
            outsider_buyout = matrix_etl.outsider_sku.buyout_pct
        elif matrix_etl.abc_group_c:
            worst = min(matrix_etl.abc_group_c, key=lambda s: s.net_profit)
            outsider_name = worst.name
            outsider_article = worst.article_id
            outsider_loss = abs(worst.net_profit)
            outsider_buyout = worst.buyout_pct
        if matrix_etl.sku_catalog:
            agg = matrix_aggregation
            sku_catalog_lines = tuple(agg.abc_a_display_lines) + tuple(
                s.catalog_line()
                for s in matrix_etl.sku_catalog[:_TOP_GROUP_A_DISPLAY]
                if _sku_label_is_valid(s.name, s.article_id)
            )
            sku_catalog_items = tuple(
                {
                    "name": s.name,
                    "article_id": s.article_id,
                    "revenue_rub": s.revenue,
                    "margin_rub": s.net_profit,
                    "buyout_pct": s.buyout_pct,
                    "abc_group": s.abc_group,
                }
                for s in matrix_etl.sku_catalog
                if _sku_label_is_valid(s.name, s.article_id)
            )
            cost_of_goods = sum(
                (d.unit_cost_rub * d.sales_qty if d.sales_qty > 0 else d.unit_cost_rub)
                for d in matrix_etl.sku_catalog
            )
            if engine is not None:
                operational_profit = engine.operational_profit
                clear_profit = engine.clear_profit
            else:
                operational_profit = (
                    revenue_total
                    - cost_of_goods
                    - storage_cost
                    - ad_cost
                    - logistics_cost
                    - commission_cost
                    - other_deductions
                    - tax
                )
                clear_profit = round(operational_profit - credit_deductions, 2)
            profitability = (
                (clear_profit / tax_base * 100.0) if tax_base > 0 else 0.0
            )
            business_score = compute_business_score(
                profitability_pct=profitability,
                ad_load_pct=adv_load,
                buyout_coef_pct=buy_ratio,
                worst_unit_net=worst_unit.net_income if worst_unit else None,
            )
            verdict = derive_business_verdict(
                business_score=business_score,
                profitability_pct=profitability,
                ad_load_pct=adv_load,
                buyout_coef_pct=buy_ratio,
                worst_unit_label=worst_label,
            )
            if clear_profit < 0 and operational_profit > 0 and credit_deductions > 0:
                verdict = (
                    "Убыток вызван удержанием по кредиту. "
                    "Операционная прибыль от продаж при этом положительна."
                )
        (
            abc_group_a_items,
            loss_sku_items,
            non_liquid_items,
            oos_zero_stock_items,
            oos_critical_stock_items,
            non_liquid_frozen_total_rub,
        ) = _collect_etl_dynamic_slices(matrix_etl)
        oos_line = normalize_finance_report_html(
            _build_oos_forecast_line(
                matrix_etl, oos_zero_stock_items, oos_critical_stock_items
            )
        )
        oos_summary = _build_oos_forecast_plain_summary(
            oos_zero_stock_items, oos_critical_stock_items
        )
        if non_liquid_frozen_total_rub > 0:
            fomo_rub = round(fomo_rub + non_liquid_frozen_total_rub, 2)

    return WbFinancePromptMetrics(
        revenue=revenue_total,
        tax=tax,
        clear_profit=clear_profit,
        adv_load_pct=adv_load,
        buy_ratio_pct=buy_ratio,
        year_forecast=revenue_total * 12,
        profitability_pct=profitability,
        business_score=business_score,
        verdict=verdict,
        fomo_lost_rub=fomo_rub,
        fomo_breakdown=fomo_parts,
        logistics_fomo_rub=logistics_fomo,
        abc_a_leader=abc_a_leader,
        abc_a_leader_name=abc_a_leader_name,
        abc_a_leader_article=abc_a_leader_article,
        abc_a_leader_revenue=abc_a_leader_revenue,
        abc_a_leader_margin=abc_a_leader_margin,
        abc_a_leader_buyout=abc_a_leader_buyout,
        abc_a_count=abc_a_count,
        abc_c_count=abc_c_count,
        abc_c_summary=abc_c_summary,
        matrix_problem_zones_block=matrix_problem_zones_block,
        localization_index_line=localization_index_line,
        outsider_name=outsider_name,
        outsider_article=outsider_article,
        outsider_loss=outsider_loss,
        outsider_buyout=outsider_buyout,
        sku_catalog_lines=sku_catalog_lines,
        sku_catalog_items=sku_catalog_items,
        oos_forecast_line=oos_line,
        oos_forecast_summary=oos_summary,
        total_ad_cost=ad_cost,
        storage_cost=storage_cost,
        credit_deductions=credit_deductions,
        operational_profit=operational_profit,
        sales_qty=wb_metrics.sales_qty if wb_metrics else 0.0,
        returns_qty=wb_metrics.returns_qty if wb_metrics else 0.0,
        deliveries_qty=wb_metrics.deliveries_qty if wb_metrics else 0.0,
        reverse_logistics_shop_avg=reverse_logistics_shop_avg,
        return_logistics_block=return_logistics_block,
        abc_group_a_items=abc_group_a_items,
        loss_sku_items=loss_sku_items,
        non_liquid_items=non_liquid_items,
        oos_zero_stock_items=oos_zero_stock_items,
        oos_critical_stock_items=oos_critical_stock_items,
        non_liquid_frozen_total_rub=non_liquid_frozen_total_rub,
        matrix_aggregation=matrix_aggregation,
    )


def _prompt_kwargs_from_metrics(metrics: WbFinancePromptMetrics) -> dict[str, str]:
    catalog_block = (
        "\n".join(f"• {line}" for line in metrics.sku_catalog_lines)
        if metrics.sku_catalog_lines
        else "—"
    )
    return {
        "revenue": _fmt_rub_in_code(metrics.revenue),
        "tax": _fmt_rub_in_code(metrics.tax),
        "clear_profit": _fmt_rub_in_code(metrics.clear_profit),
        "profitability_pct": f"{metrics.profitability_pct:.1f}",
        "adv_load": f"{metrics.adv_load_pct:.1f}",
        "buy_ratio": f"{metrics.buy_ratio_pct:.1f}",
        "year_forecast": _fmt_rub_in_code(metrics.year_forecast, decimals=0),
        "business_score": f"{metrics.business_score:.1f}",
        "verdict": metrics.verdict,
        "fomo_lost_rub": _fmt_rub_in_code(metrics.fomo_lost_rub),
        "logistics_fomo_rub": _fmt_rub_in_code(metrics.logistics_fomo_rub),
        "abc_a_leader": metrics.abc_a_leader,
        "abc_a_leader_name": metrics.abc_a_leader_name,
        "abc_a_leader_article": metrics.abc_a_leader_article,
        "abc_a_count": str(metrics.abc_a_count),
        "abc_c_count": str(metrics.abc_c_count),
        "abc_c_summary": metrics.abc_c_summary,
        "matrix_problem_zones_block": metrics.matrix_problem_zones_block,
        "outsider_name": metrics.outsider_name,
        "outsider_article": metrics.outsider_article,
        "outsider_loss": _fmt_rub_in_code(metrics.outsider_loss),
        "outsider_buyout": f"{metrics.outsider_buyout:.1f}",
        "abc_a_leader_buyout": f"{metrics.abc_a_leader_buyout:.1f}",
        "sku_catalog_block": catalog_block,
        "fomo_details_block": _format_fomo_details_block(metrics.fomo_breakdown),
        "return_logistics_block": metrics.return_logistics_block,
        "reverse_logistics_avg_rub": f"{metrics.reverse_logistics_shop_avg:.2f}",
        "oos_forecast_line": metrics.oos_forecast_line,
        "localization_index_line": metrics.localization_index_line,
    }


def build_wb_marketplace_finance_payload_dict(
    prompt_metrics: WbFinancePromptMetrics,
    wb_metrics: WbMarketplaceMetrics | None,
) -> dict[str, Any]:
    """JSON-словарь user-сообщения для OpenRouter (wb_ozon_finance)."""
    top5: list[dict[str, Any]] = []
    traffic_light: dict[str, str] = {}
    if wb_metrics:
        top5 = [
            {
                "label": row.label,
                "sale_price": round(row.sale_price, 2),
                "unit_logistics": round(row.unit_logistics, 2),
                "net_income_per_unit": round(row.net_income, 2),
            }
            for row in wb_metrics.top5_units
        ]
        if wb_metrics.top5_units:
            best = max(wb_metrics.top5_units, key=lambda u: u.net_income)
            worst = min(wb_metrics.top5_units, key=lambda u: u.net_income)
            traffic_light["green_scale"] = (
                f"Лидер «{best.label}» — чистый доход {best.net_income:.2f} руб./шт."
            )
            if wb_metrics.ad_load_pct <= 15:
                traffic_light["green_ad"] = f"ДРР {wb_metrics.ad_load_pct:.1f}% под контролем."
            if 45 <= wb_metrics.buyout_coef_pct < 65:
                traffic_light["yellow_buyout"] = (
                    f"Выкуп {wb_metrics.buyout_coef_pct:.1f}% — на грани, нужна доработка карточек."
                )
            elif 10 < wb_metrics.ad_load_pct <= 22:
                traffic_light["yellow_ad"] = (
                    f"Реклама {wb_metrics.ad_load_pct:.1f}% — следите за окупаемостью кампаний."
                )
            if worst.net_income < 0:
                traffic_light["red_sku"] = (
                    f"«{worst.label}» убыточен: {worst.net_income:.2f} руб./шт. — вымывает оборотку."
                )
            if wb_metrics.buyout_coef_pct > 0 and wb_metrics.buyout_coef_pct < 45:
                traffic_light["red_buyout"] = (
                    f"Выкуп {wb_metrics.buyout_coef_pct:.1f}% — критическая зона возвратов."
                )

    return {
        "etl_source": "wb_ozon_finance_report",
        "revenue_rub": round(prompt_metrics.revenue, 2),
        "tax_usn_6pct_rub": round(prompt_metrics.tax, 2),
        "clear_profit_rub": round(prompt_metrics.clear_profit, 2),
        "profitability_pct": round(prompt_metrics.profitability_pct, 1),
        "business_score": prompt_metrics.business_score,
        "business_verdict": prompt_metrics.verdict,
        "fomo_lost_rub": round(prompt_metrics.fomo_lost_rub, 2),
        "fomo_breakdown": list(prompt_metrics.fomo_breakdown),
        "logistics_fomo_rub": round(prompt_metrics.logistics_fomo_rub, 2),
        "abc_analysis": {
            "group_a_leader_name": prompt_metrics.abc_a_leader_name,
            "group_a_leader_article": prompt_metrics.abc_a_leader_article,
            "group_a_leader_revenue_rub": round(prompt_metrics.abc_a_leader_revenue, 2),
            "group_a_leader_margin_rub": round(prompt_metrics.abc_a_leader_margin, 2),
            "group_a_leader_buyout_pct": round(prompt_metrics.abc_a_leader_buyout, 1),
            "group_a_count": prompt_metrics.abc_a_count,
            "group_c_count": prompt_metrics.abc_c_count,
            "group_c_summary": prompt_metrics.abc_c_summary,
            "matrix_problem_zones": prompt_metrics.matrix_problem_zones_block,
            "localization_index": prompt_metrics.localization_index_line,
        },
        "outsider_sku": {
            "name": prompt_metrics.outsider_name,
            "article_id": prompt_metrics.outsider_article,
            "loss_rub": round(prompt_metrics.outsider_loss, 2),
            "buyout_pct": round(prompt_metrics.outsider_buyout, 1),
        },
        "sku_catalog": list(prompt_metrics.sku_catalog_items),
        "out_of_stock_forecast": prompt_metrics.oos_forecast_line,
        "ad_load_pct": round(prompt_metrics.adv_load_pct, 1),
        "buyout_coef_pct": round(prompt_metrics.buy_ratio_pct, 1),
        "year_forecast_rub": round(prompt_metrics.year_forecast, 0),
        "total_advertising_cost_rub": round(prompt_metrics.total_ad_cost, 2),
        "sales_qty": prompt_metrics.sales_qty,
        "returns_qty": prompt_metrics.returns_qty,
        "deliveries_qty": prompt_metrics.deliveries_qty,
        "top5_unit_economics": top5,
        "traffic_light_hints": traffic_light,
        "local_insights": list(wb_metrics.insight_lines) if wb_metrics else [],
    }


def build_wb_marketplace_finance_user_prompt(
    prompt_metrics: WbFinancePromptMetrics,
    wb_metrics: WbMarketplaceMetrics | None,
    *,
    matrix_rows: list[list[str]] | None = None,
) -> str:
    """User-сообщение: MPSTATS JSON (все метрики из Python ETL)."""
    ctx = build_wb_mpstats_ai_context(
        matrix_rows or [],
        revenue_total=prompt_metrics.revenue,
        platform="wildberries",
    )
    if ctx.get("error"):
        ctx = build_wb_marketplace_finance_payload_dict(prompt_metrics, wb_metrics)
    return build_wb_finance_json_user_message(ctx)


def build_wb_finance_system_prompt_from_totals(
    revenue_total: float,
    wb_metrics: WbMarketplaceMetrics | None = None,
    *,
    matrix_rows: list[list[str]] | None = None,
) -> str | None:
    """Удобная обёртка: revenue + wb_metrics → готовый system prompt."""
    metrics = compute_wb_finance_prompt_metrics(
        revenue_total, wb_metrics, matrix_rows=matrix_rows
    )
    if metrics is None:
        return None
    return build_wb_marketplace_finance_system_prompt(**_prompt_kwargs_from_metrics(metrics))


def build_wb_finance_openrouter_prompt_pair(
    matrix_rows: list[list[str]],
    *,
    revenue_total: float,
    wb_metrics: WbMarketplaceMetrics | None = None,
    file_path: str | Path | None = None,
    platform: str | None = "wildberries",
) -> tuple[str, str] | None:
    """Пара system + user для OpenRouter: cfo-v8 + MPSTATS JSON из Python."""
    from content.chat_prompt import WB_ANALYTICS_SYSTEM_PROMPT

    ctx = resolve_wb_mpstats_context(
        file_path=file_path,
        matrix_rows=matrix_rows,
        revenue_total=revenue_total,
        platform=platform,
    )
    if ctx.get("error"):
        return None
    system = WB_ANALYTICS_SYSTEM_PROMPT
    user = build_wb_finance_json_user_message(ctx)
    return system, user


def has_legacy_wb_finance_markers(text: str) -> bool:
    """True, если в тексте остались устаревшие маркеры CFO-отчёта."""
    sample = (text or "").strip()
    if not sample:
        return True
    return any(marker.lower() in sample.lower() for marker in _LEGACY_FINANCE_MARKERS)


def sanitize_wb_finance_html(html: str) -> str:
    """Нормализует устаревшие заголовки и формулировки в ответе модели."""
    text = normalize_finance_report_html((html or "").strip())
    for pattern, replacement in _LEGACY_FINANCE_REPLACEMENTS:
        text = pattern.sub(replacement, text)
    text = _TECH_HINT_PAREN_RE.sub("", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _abc_sku_to_dict(sku: object, group: str) -> dict[str, Any]:
    """MatrixAbcSku / MatrixSkuDetail → JSON для Mini App."""
    name = getattr(sku, "name", "—")
    article = getattr(sku, "article_id", "") or name
    revenue = float(getattr(sku, "revenue", 0) or 0)
    margin = float(getattr(sku, "net_profit", getattr(sku, "margin", 0)) or 0)
    buyout = float(getattr(sku, "buyout_pct", 0) or 0)
    return {
        "sku": str(article),
        "name": str(name),
        "article_id": str(article),
        "revenue": round(revenue, 2),
        "revenue_rub": round(revenue, 2),
        "margin": round(margin, 2),
        "margin_rub": round(margin, 2),
        "buyout_pct": round(buyout, 1),
        "abc_group": group,
    }


_STRIP_HTML_TAGS_RE = re.compile(r"<[^>]+>")


def _strip_html_tags(text: str) -> str:
    return _STRIP_HTML_TAGS_RE.sub("", text or "").replace("&nbsp;", " ").strip()


def _matrix_aggregation_to_dict(agg: WbFinanceMatrixAggregation) -> dict[str, Any]:
    return {
        "tail_a_count": agg.tail_a_count,
        "tail_c_count": agg.tail_c_count,
        "tail_c_revenue_rub": round(agg.tail_c_revenue, 2),
        "tail_ballast_count": agg.tail_ballast_count,
        "tail_ballast_loss_rub": round(agg.tail_ballast_loss, 2),
        "tail_frozen_count": agg.tail_frozen_count,
        "tail_frozen_stock": agg.tail_frozen_stock,
        "tail_loss_sku_count": agg.tail_loss_sku_count,
        "tail_loss_sku_rub": round(agg.tail_loss_sku_rub, 2),
        "abc_a_display": list(agg.abc_a_display_lines),
        "abc_c_display": list(agg.abc_c_display_lines),
        "ballast_display": list(agg.ballast_display_lines),
        "non_liquid_display": list(agg.non_liquid_display_lines),
    }


def _build_mini_app_traffic_light_dict(
    prompt_metrics: WbFinancePromptMetrics,
    wb_metrics: WbMarketplaceMetrics | None,
) -> dict[str, str]:
    """Плоский светофор для Mini App (без HTML-тегов)."""
    block = _build_traffic_light_block(wb_metrics, prompt_metrics)
    keys = ("green", "yellow", "red")
    out: dict[str, str] = {}
    for idx, key in enumerate(keys):
        raw = block[idx * 2] if idx * 2 < len(block) else ""
        text = _strip_html_tags(raw)
        for prefix in (
            "ЗОНА УСПЕХА:",
            "ЗОНА ВНИМАНИЯ:",
            "КРИТИЧЕСКАЯ ЗОНА:",
        ):
            text = text.replace(prefix, "").strip()
        out[key] = text or (
            "Лидеры с высокой маржой в этом периоде отсутствуют"
            if key == "green"
            else "показатели в норме"
            if key == "yellow"
            else "Критических убытков по товарам не зафиксировано"
        )
    return out


def _mini_app_abc_groups_from_etl(
    matrix_etl: object | None,
    agg: WbFinanceMatrixAggregation,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """ТОП-N SKU для Mini App; полная математика остаётся в summary."""
    catalog_map: dict[str, Any] = {}
    if matrix_etl is not None:
        catalog_map = {s.name: s for s in getattr(matrix_etl, "sku_catalog", ()) or ()}

    group_a: list[dict[str, Any]] = []
    for item in agg.abc_group_a_display_items:
        detail = catalog_map.get(str(item.get("name", "")))
        revenue = float(detail.revenue) if detail else 0.0
        buyout = float(item.get("buyout_pct") or (detail.buyout_pct if detail else 0.0))
        group_a.append(
            {
                "sku": str(item.get("article_id") or item.get("name") or "—"),
                "name": str(item.get("name") or "—"),
                "article_id": str(item.get("article_id") or "—"),
                "revenue": round(revenue, 2),
                "revenue_rub": round(revenue, 2),
                "margin": round(float(item.get("net_profit_rub", 0.0)), 2),
                "margin_rub": round(float(item.get("net_profit_rub", 0.0)), 2),
                "unit_profit_rub": round(float(item.get("unit_profit_rub", 0.0)), 2),
                "buyout_pct": round(buyout, 1),
                "abc_group": "A",
            }
        )

    group_b: list[dict[str, Any]] = []
    group_c: list[dict[str, Any]] = []
    if matrix_etl is not None:
        all_c = sorted(
            getattr(matrix_etl, "abc_group_c", ()) or (),
            key=lambda s: float(s.revenue),
        )
        for sku in all_c[:_TOP_GROUP_C_DISPLAY]:
            group_c.append(_abc_sku_to_dict(sku, "C"))

        ranked_b = [
            s
            for s in getattr(matrix_etl, "sku_catalog", ()) or ()
            if (s.abc_group or "B").upper() == "B"
        ]
        ranked_b.sort(key=lambda s: float(s.net_profit), reverse=True)
        for item in ranked_b[:_TOP_GROUP_A_DISPLAY]:
            group_b.append(_abc_sku_to_dict(item, "B"))

    return group_a, group_b, group_c


def _mini_app_sku_catalog_slice(
    group_a: list[dict[str, Any]],
    group_b: list[dict[str, Any]],
    group_c: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Компактный каталог для WebView (без 10k карточек)."""
    seen: set[str] = set()
    out: list[dict[str, Any]] = []
    for batch in (group_a, group_b, group_c):
        for row in batch:
            key = str(row.get("article_id") or row.get("sku") or row.get("name"))
            if key in seen:
                continue
            seen.add(key)
            out.append(row)
    return out


def build_wb_finance_mini_app_extensions(
    revenue_total: float,
    wb_metrics: WbMarketplaceMetrics | None,
    *,
    matrix_rows: list[list[str]] | None,
    platform: str | None = None,
) -> dict[str, Any] | None:
    """Расширения table_raw_json для Mini App (ABC, SKU, summary)."""
    from services.marketplace_platform import normalize_marketplace_platform

    platform_id = normalize_marketplace_platform(platform)
    prompt_metrics = compute_wb_finance_prompt_metrics(
        revenue_total, wb_metrics, matrix_rows=matrix_rows, platform=platform_id
    )
    if prompt_metrics is None:
        return None

    from services.file_processor import compute_seller_matrix_etl

    matrix_etl = (
        compute_seller_matrix_etl(
            matrix_rows,
            revenue_total=revenue_total,
            platform=platform_id,
        )
        if matrix_rows
        else None
    )
    agg = prompt_metrics.matrix_aggregation
    if matrix_etl is not None and not agg.abc_a_display_lines:
        agg = aggregate_matrix_display(matrix_etl)

    group_a, group_b, group_c = _mini_app_abc_groups_from_etl(matrix_etl, agg)

    oos_forecast: list[dict[str, Any]] = []
    if matrix_etl:
        for oos in matrix_etl.oos_forecasts:
            fomo = 0.0
            if oos.risk_out_of_stock and oos.days_until_stockout is not None:
                shortage = max(0.0, 7.0 - oos.days_until_stockout)
                fomo = max(0.0, shortage * 500.0)
            oos_forecast.append({
                "sku": oos.label,
                "name": oos.label,
                "days_until_stockout": (
                    round(oos.days_until_stockout, 1)
                    if oos.days_until_stockout is not None
                    else None
                ),
                "risk_out_of_stock": oos.risk_out_of_stock,
                "fomo_lost_rub": round(fomo, 2),
            })

    score_emoji, score_status = _business_score_band(prompt_metrics.business_score)
    sku_catalog = _mini_app_sku_catalog_slice(group_a, group_b, group_c)

    return {
        "source": "wb_ozon_finance_xlsx",
        "platform": platform_id,
        "cfo_build": _FINANCE_REPORT_BUILD,
        "abc_analysis": {
            "group_a": group_a,
            "group_b": group_b,
            "group_c": group_c,
        },
        "matrix_aggregation": _matrix_aggregation_to_dict(agg),
        "sku_catalog": sku_catalog,
        "out_of_stock_forecast": oos_forecast,
        "health_index": {
            "score": round(prompt_metrics.business_score, 1),
            "emoji": score_emoji,
            "status": score_status,
            "verdict": prompt_metrics.verdict,
        },
        "traffic_light": _build_mini_app_traffic_light_dict(prompt_metrics, wb_metrics),
        "loss_calculator": {
            "fomo_lost_rub": round(prompt_metrics.fomo_lost_rub, 2),
            "non_liquid_frozen_rub": round(prompt_metrics.non_liquid_frozen_total_rub, 2),
            "return_logistics_shop_avg_rub": round(prompt_metrics.reverse_logistics_shop_avg, 2),
        },
        "summary": {
            "revenue": round(prompt_metrics.revenue, 2),
            "revenue_rub": round(prompt_metrics.revenue, 2),
            "net_profit": round(prompt_metrics.clear_profit, 2),
            "business_score": prompt_metrics.business_score,
            "profitability_pct": round(prompt_metrics.profitability_pct, 1),
            "ad_load_pct": round(prompt_metrics.adv_load_pct, 1),
            "buyout_coef_pct": round(prompt_metrics.buy_ratio_pct, 1),
            "fomo_rub": round(prompt_metrics.fomo_lost_rub, 2),
            "group_a_leader": prompt_metrics.abc_a_leader_name,
            "reverse_logistics_shop_avg": round(prompt_metrics.reverse_logistics_shop_avg, 2),
        },
    }


def enrich_table_json_wb_finance(
    table_json: str,
    *,
    revenue_total: float,
    wb_metrics: WbMarketplaceMetrics | None = None,
    matrix_rows: list[list[str]] | None = None,
    platform: str | None = None,
) -> str:
    """Дополняет канонический JSON отчёта полями для Mini App дашборда."""
    from services.table_json import canonicalize_table_json

    extensions = build_wb_finance_mini_app_extensions(
        revenue_total,
        wb_metrics,
        matrix_rows=matrix_rows,
        platform=platform,
    )
    if not extensions:
        return table_json
    try:
        payload = json.loads(table_json)
    except json.JSONDecodeError:
        return table_json
    if not isinstance(payload, dict):
        return table_json
    payload.update(extensions)
    return canonicalize_table_json(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    )


def append_wb_finance_mini_app_cta(html: str) -> str:
    """Раньше добавлял CTA в чат; Studio открывается через MenuButtonWebApp."""
    return repair_telegram_html((html or "").strip())


def _format_abc_a_leader_html(prompt_metrics: WbFinancePromptMetrics) -> list[str]:
    """Группа A: ТОП-5 лидеров + агрегированный хвост."""
    agg = prompt_metrics.matrix_aggregation
    if prompt_metrics.revenue > 0 and (
        not agg.abc_a_display_lines
        and prompt_metrics.abc_a_count == 0
    ):
        return [
            (
                "🅰️ <b>Товары-лидеры (Приносят основные деньги группы А):</b> "
                "<i>Отсутствуют, выручка требует оптимизации</i>"
            )
        ]
    lines = ["🅰️ <b>Товары-лидеры (Приносят основные деньги группы А):</b>"]
    if agg.abc_a_display_lines:
        lines.extend(agg.abc_a_display_lines)
    elif prompt_metrics.abc_a_leader_name not in ("—", ""):
        label = _escape_verdict(
            _format_sku_label(
                prompt_metrics.abc_a_leader_name,
                prompt_metrics.abc_a_leader_article,
            )
        )
        lines.append(
            f"• <b>{label}</b> — выкуп "
            f"<code>{prompt_metrics.abc_a_leader_buyout:.1f}%</code>"
        )
    if agg.tail_a_count > 0:
        lines.append(_tail_line_group_a(agg.tail_a_count))
    return lines


def _cfo_metrics_health_band(
    margin_pct: float,
    *,
    total_system_losses: float,
    total_storage_cost: float,
) -> tuple[float, str, str, str]:
    """Индекс здоровья магазина из CFO-метрик (1–10, эмодзи, вывод, статус)."""
    score = 10.0
    emoji = "🟢"
    conclusion = "Высокая маржинальность — фокус на масштабировании лидеров ассортимента."
    if margin_pct < 10:
        score -= 4.0
        emoji = "🔴"
        conclusion = (
            "Критически низкая рентабельность — угроза кассового разрыва в ближайшем цикле."
        )
    elif margin_pct < 15:
        score -= 1.0
        emoji = "🟡"
        conclusion = "Рентабельность ниже целевой — пересмотрите рекламу и складские издержки."
    if total_system_losses > 1000 or total_storage_cost > 2000:
        score -= 1.5
    score = max(1.0, min(10.0, score))
    if score >= 8.0:
        status = "ВЫСОКИЙ УРОВЕНЬ"
    elif score >= 5.0:
        status = "СРЕДНИЙ УРОВЕНЬ"
    else:
        status = "КРИТИЧЕСКИЙ УРОВЕНЬ"
    return score, emoji, conclusion, status


def _pick_cfo_top_sku(
    sku_data: dict[str, dict[str, float | int]],
) -> tuple[str, float, float]:
    """Лидер группы A: подпись с артикулом, выкуп %, чистая прибыль со штуки."""
    if not sku_data:
        return "Товары не найдены", 0.0, 0.0
    top_sku = max(
        sku_data,
        key=lambda sku: float(sku_data[sku].get("rrc_revenue", 0.0)),
    )
    stats = sku_data[top_sku]
    label = _sku_display_from_stats(top_sku, stats)
    sales = int(stats.get("sales_count", 0))
    returns = int(stats.get("returns_count", 0))
    denom = sales + returns
    buyout = (sales / denom * 100.0) if denom > 0 else 100.0
    unit_profit = float(stats.get("unit_profit_rub", 0.0) or 0.0)
    if unit_profit == 0.0 and sales > 0:
        net = float(stats.get("net_profit_rub", 0.0) or 0.0)
        unit_profit = round(net / sales, 2)
    return label, buyout, unit_profit


def _build_storage_traffic_text(storage_cost: float, *, html: bool = False) -> str:
    if storage_cost <= 0:
        text = (
            "🟢 Издержки на хранение отсутствуют (0.00 руб.). "
            "Оборачиваемость идеальная, замороженного капитала нет."
        )
    else:
        amount = _fmt_cfo_pre_money(storage_cost)
        text = (
            f"🟡 ЗОНА ВНИМАНИЯ: Списания за хранение: {amount} руб. "
            "— пересмотрите оборачиваемость и объём неликвида на складе."
        )
    if html:
        return text.replace(
            _fmt_cfo_pre_money(storage_cost),
            f"<code>{_fmt_rub_in_code(storage_cost)}</code>",
            1,
        ) if storage_cost > 0 else text
    return text


def _build_penalties_traffic_text(
    system_losses: float,
    credit_deductions: float = 0.0,
    *,
    html: bool = False,
) -> str:
    total = round(float(system_losses) + float(credit_deductions), 2)
    if total <= 0:
        return (
            "🟢 Системных штрафов, удержаний по кредитам и санкций за маркировку от WB "
            "не зафиксировано (0.00 руб.). Карточки заполнены верно."
        )
    amount = _fmt_cfo_pre_money(total)
    text = f"🔴 КРИТИЧЕСКАЯ ЗОНА: Штрафы и удержания WB: {amount} руб."
    if html:
        return (
            "🔴 <b>КРИТИЧЕСКАЯ ЗОНА:</b> Штрафы и удержания WB: "
            f"<code>{_fmt_rub_in_code(total)}</code> руб."
        )
    return text


def _build_loss_calculator_headline(fomo_lost_rub: float, *, html: bool = False) -> str:
    if fomo_lost_rub <= 0:
        return (
            "🛡️ Эффективность юнит-экономики 100%. Скрытых операционных потерь на логистике "
            "возвратов не выявлено. Вся маржа идет в карман."
        )
    amount_plain = _fmt_cfo_pre_money(fomo_lost_rub)
    plain = f"Потенциально можно вернуть в оборот: {amount_plain} руб."
    if html:
        return (
            f"Потенциально можно вернуть в оборот: "
            f"<code>{_fmt_rub_in_code(fomo_lost_rub)}</code> руб."
        )
    return plain


def _fmt_cfo_pre_money(value: float) -> str:
    return f"{float(value):,.2f}".replace(",", " ")


def build_wb_finance_consulting_html_from_cfo_metrics(metrics: dict[str, object]) -> str:
    """
    Генерирует финальный HTML-отчёт CFO Engine v12 из ``sync_table_cfo_processing_worker``.
    """
    if metrics.get("error"):
        return (
            f"<pre>❌ Не удалось построить CFO-отчёт: "
            f"{html.escape(str(metrics.get('error')))}</pre>"
        )

    margin_pct = float(metrics.get("margin_pct", 0.0) or 0.0)
    total_revenue = float(metrics.get("total_revenue", 0.0) or 0.0)
    tax_total = float(metrics.get("tax_total", 0.0) or 0.0)
    net_profit = float(metrics.get("net_profit", 0.0) or 0.0)
    storage = float(metrics.get("total_storage_cost", 0.0) or 0.0)
    system_losses = float(metrics.get("total_system_losses", 0.0) or 0.0)
    tax_type = str(metrics.get("tax_type") or "USN")
    tax_rate = float(metrics.get("tax_rate", 6.0) or 0.0)
    sku_data: dict[str, dict[str, Any]] = dict(metrics.get("sku_data") or {})

    oos_zero_norm = [
        _normalize_oos_stock_item(item, sku_data)
        for item in list(metrics.get("oos_zero_stock_items") or [])
    ]
    oos_critical_norm = [
        _normalize_oos_stock_item(item, sku_data)
        for item in list(metrics.get("oos_critical_sku") or [])
    ]

    health_score, health_emoji, conclusion, health_status = _cfo_metrics_health_band(
        margin_pct,
        total_system_losses=system_losses,
        total_storage_cost=storage,
    )

    if tax_type == "NONE":
        tax_string = "📉 НАЛОГ: Не учитывается (0.00 руб.)"
    else:
        tax_string = (
            f"📉 НАЛОГ {tax_type} ({tax_rate:.0f}%): "
            f"{_fmt_cfo_pre_money(tax_total)} руб."
        )

    oos_pre = normalize_finance_report_html(
        _build_oos_forecast_line(
            None,
            tuple(oos_zero_norm),
            tuple(oos_critical_norm),
        )
    )
    oos_inner = oos_pre.removeprefix("<pre>").removesuffix("</pre>")

    top_label, top_buyout, top_unit_profit = _pick_cfo_top_sku(sku_data)
    top_label_esc = html.escape(top_label)

    abc_lines: list[str] = []
    ranked = sorted(
        sku_data.items(),
        key=lambda kv: float(kv[1].get("rrc_revenue", 0.0) or 0.0),
        reverse=True,
    )
    for sku_key, stats in ranked[:7]:
        revenue = float(stats.get("rrc_revenue", 0.0) or 0.0)
        if revenue <= 0:
            continue
        label = html.escape(_sku_display_from_stats(sku_key, stats))
        sales = int(stats.get("sales_count", 0))
        returns = int(stats.get("returns_count", 0))
        denom = sales + returns
        buyout = (sales / denom * 100.0) if denom > 0 else 0.0
        unit_profit = float(stats.get("unit_profit_rub", 0.0) or 0.0)
        abc_lines.append(
            f"• {label} — чистая прибыль {_fmt_cfo_pre_money(unit_profit)} руб./шт., "
            f"выкуп {buyout:.1f}%"
        )
    abc_block = "\n".join(abc_lines) if abc_lines else "• лидеры не выявлены"

    outsider_lines: list[str] = []
    for sku_key, stats in sorted(
        sku_data.items(),
        key=lambda kv: float(kv[1].get("net_profit_rub", 0.0) or 0.0),
    ):
        net = float(stats.get("net_profit_rub", 0.0) or 0.0)
        if net >= 0:
            continue
        label = html.escape(_sku_display_from_stats(sku_key, stats))
        outsider_lines.append(
            f"• {label} — убыток {_fmt_cfo_pre_money(abs(net))} руб."
        )
    outsider_block = (
        "\n".join(outsider_lines[:5])
        if outsider_lines
        else "убыточных товаров не выявлено"
    )

    green_zone = (
        f"🟢 ЗОНА УСПЕХА: Масштабируйте {top_label_esc} "
        f"(чистая прибыль {_fmt_cfo_pre_money(top_unit_profit)} руб./шт.) "
        "при наличии остатков на складе отгрузки."
    )
    if storage <= 0:
        green_zone += f"\n{_build_storage_traffic_text(0.0)}"

    yellow_zone = (
        _build_storage_traffic_text(storage)
        if storage > 0
        else "🟡 ЗОНА ВНИМАНИЯ: критических операционных отклонений не зафиксировано."
    )
    red_zone = _build_penalties_traffic_text(system_losses)

    loss_pool = round(storage + system_losses, 2)
    loss_headline = _build_loss_calculator_headline(loss_pool)
    loss_details: list[str] = []
    if storage > 0:
        loss_details.append(
            f"• Складское хранение: {_fmt_cfo_pre_money(storage)} руб."
        )
    if system_losses > 0:
        loss_details.append(
            f"• Системные штрафы и удержания: {_fmt_cfo_pre_money(system_losses)} руб."
        )
    loss_details_block = "\n".join(loss_details)

    zero_labels = {_oos_item_display_label(z) for z in oos_zero_norm}
    if top_label in zero_labels:
        action_1 = (
            f"Срочно закупите лидера {top_label_esc} — товар закончился на складе. "
            f"Целевая чистая прибыль со штуки: {_fmt_cfo_pre_money(top_unit_profit)} руб."
        )
    else:
        action_1 = (
            f"Усилить закуп и рекламу на {top_label_esc} — "
            f"чистая прибыль с одной продажи: {_fmt_cfo_pre_money(top_unit_profit)} руб."
        )

    if oos_zero_norm:
        names = ", ".join(
            html.escape(_oos_item_display_label(z)) for z in oos_zero_norm[:3]
        )
        action_3 = (
            f"Срочно закупите: {names} — "
            "товар закончился на складе при сохранённом спросе."
        )
    elif oos_critical_norm:
        names = ", ".join(
            html.escape(_oos_item_display_label(c)) for c in oos_critical_norm[:3]
        )
        action_3 = (
            f"Запланируйте поставку в течение 48 часов: {names} — "
            "остаток на складе критически мал при текущем темпе продаж."
        )
    else:
        action_3 = (
            "Контролируйте остатки по рисковым артикулам: "
            "критических рисков обнуления остатков не выявлено."
        )

    year_forecast = total_revenue * 12.0

    report_html = f"""<pre>
📊 ФИНАНСОВЫЙ ЭКСПРЕСС-АНАЛИЗ МАГАЗИНА

────────────────────────

🎯 ИНДЕКС ЗДОРОВЬЯ БИЗНЕСА: {health_emoji} {health_score:.1f} / 10 {health_status}
📉 Балл рассчитан автоматически на основе рентабельности и издержек хранения.

💡 ГЛАВНЫЙ АНАЛИТИЧЕСКИЙ ВЫВОД:
{html.escape(conclusion)}

────────────────────────

💰 ОБЩАЯ ВЫРУЧКА: {_fmt_cfo_pre_money(total_revenue)} руб.
{tax_string}
💵 ЧИСТАЯ ПРИБЫЛЬ: {_fmt_cfo_pre_money(net_profit)} руб.

Эффективность (рентабельность) чистой прибыли: {margin_pct:.1f}%

────────────────────────

📦 ABC-АНАЛИЗ ПРОДАЖ
🅰️ Товары-лидеры (Приносят основные деньги группы А):
{abc_block}

🅲 Товары-аутсайдеры (Слабые продажи или убытки):
{outsider_block}

────────────────────────

📈 СВЕТОФОР ЭФФЕКТИВНОСТИ
{green_zone}
{yellow_zone}
{red_zone}

────────────────────────

💸 КАЛЬКУЛЯТОР ПОТЕРЬ И УПУЩЕННОЙ ВЫГОДЫ
{loss_headline}
{loss_details_block}

────────────────────────

🛡️ ПРОГНОЗ И ОБНУЛЕНИЕ ОСТАТКОВ
При сохранении текущего темпа годовой оборот составит около {_fmt_cfo_pre_money(year_forecast)} руб.
{oos_inner}

────────────────────────

📋 ПЛАН ДЕЙСТВИЙ ДЛЯ ПРЕДПРИНИМАТЕЛЯ НА СЕГОДНЯ
1. {action_1}
2. Держите ДРР не выше 15-20% — еженедельно чистите неокупаемые поисковые ключи в кампаниях.
3. {action_3}

CFO build {CFO_BUILD_FOOTER_PLAIN}</pre>
"""
    return normalize_finance_report_html(
        append_wb_finance_mini_app_cta(report_html)
    )


def build_wb_finance_express_html_local(
    prompt_metrics: WbFinancePromptMetrics,
    wb_metrics: WbMarketplaceMetrics | None,
) -> str:
    """Re-export: каноническая сборка отчёта — :func:`services.table_text_response.build_wb_finance_express_html_local`."""
    from services.table_text_response import build_wb_finance_express_html_local as _impl

    return _impl(prompt_metrics, wb_metrics)


def _build_loss_calculator_lines(prompt_metrics: WbFinancePromptMetrics) -> list[str]:
    """Калькулятор потерь cfo-v12 — без шаблонов при нулевых суммах."""
    lines = [
        _FINANCE_SEPARATOR,
        "💸 <b>КАЛЬКУЛЯТОР ПОТЕРЬ И УПУЩЕННОЙ ВЫГОДЫ</b>",
        _build_loss_calculator_headline(prompt_metrics.fomo_lost_rub, html=True),
    ]

    if prompt_metrics.fomo_lost_rub <= 0:
        return lines

    detail_lines: list[str] = []
    for part in _expand_fomo_breakdown(prompt_metrics.fomo_breakdown):
        low = part.lower()
        if "неликвид" in low:
            continue
        detail_lines.append(f"• {_escape_verdict(part)}")

    for item in prompt_metrics.matrix_aggregation.non_liquid_display_items:
        stock = int(item.get("stock", 0))
        if stock <= 0:
            continue
        label = _escape_verdict(str(item.get("sku", item.get("name", "—"))))
        unit_cost = float(item.get("cost", 0.0))
        frozen = float(item.get("frozen_capital_rub", 0.0))
        if unit_cost > 0 and frozen > 0:
            detail_lines.append(
                "• "
                f"<b>{label}</b> — остаток <code>{stock}</code> шт. × "
                f"себестоимость <code>{_fmt_rub_in_code(unit_cost)}</code> = "
                f"заморожено <code>{_fmt_rub_in_code(frozen)}</code> руб."
            )
        else:
            detail_lines.append(
                "• "
                f"<b>{label}</b> — остаток <code>{stock}</code> шт., "
                "движения в периоде нет (себестоимость не указана в отчёте)."
            )
    agg = prompt_metrics.matrix_aggregation
    if agg.tail_frozen_count > 0:
        detail_lines.append(
            _escape_verdict(
                _tail_line_non_liquid(agg.tail_frozen_count, agg.tail_frozen_stock)
            )
        )

    if detail_lines:
        lines.extend(detail_lines)

    if prompt_metrics.non_liquid_frozen_total_rub > 0:
        lines.append(
            f"Заморожено в неликвиде: "
            f"<code>{_fmt_rub_in_code(prompt_metrics.non_liquid_frozen_total_rub)} руб.</code>"
        )

    lines.append(
        f"<i>Исправление выявленных зон вернёт в оборот до "
        f"<code>{_fmt_rub_in_code(prompt_metrics.fomo_lost_rub)} руб.</code></i>"
    )
    return lines


def _build_traffic_light_block(
    wb_metrics: WbMarketplaceMetrics | None,
    prompt_metrics: WbFinancePromptMetrics,
) -> list[str]:
    """🟢🟡🔴 cfo-v12 — только факты из ETL, без пугающих шаблонов при 0.00 руб."""
    agg = prompt_metrics.matrix_aggregation
    storage = float(prompt_metrics.storage_cost or 0.0)
    credit = float(prompt_metrics.credit_deductions or 0.0)

    green = "🟢 <b>ЗОНА УСПЕХА:</b> "
    display_a = agg.abc_group_a_display_items or prompt_metrics.abc_group_a_items[:_TOP_GROUP_A_DISPLAY]
    if display_a:
        green_parts = [
            (
                f"<b>{_escape_verdict(_oos_item_display_label(dict(item)))}</b> — "
                f"чистая прибыль <code>{_fmt_rub_in_code(float(item.get('unit_profit_rub', 0.0)))}</code>/шт."
            )
            for item in display_a
        ]
        green += "; ".join(green_parts)
        if agg.tail_a_count > 0:
            green += f" {_escape_verdict(_tail_line_group_a(agg.tail_a_count).removeprefix('• '))}"
    else:
        green += "Лидеры с высокой маржой в этом периоде отсутствуют"
    if storage <= 0:
        green += f" {_build_storage_traffic_text(0.0)}"

    yellow = "🟡 <b>ЗОНА ВНИМАНИЯ:</b> "
    yellow_parts: list[str] = []
    if storage > 0:
        yellow_parts.append(
            f"Списания за хранение: <code>{_fmt_rub_in_code(storage)}</code> руб. "
            "— пересмотрите оборачиваемость и объём неликвида."
        )
    if prompt_metrics.total_ad_cost > 0 and prompt_metrics.adv_load_pct >= _DRR_WARNING_PCT:
        yellow_parts.append(
            f"ДРР <code>{prompt_metrics.adv_load_pct:.1f}%</code> — "
            f"рекламные расходы <code>{_fmt_rub_in_code(prompt_metrics.total_ad_cost)}</code> руб."
        )
    elif wb_metrics and prompt_metrics.total_ad_cost > 0 and 12 < wb_metrics.ad_load_pct <= _HIGH_DRR_PCT:
        yellow_parts.append(
            f"реклама <code>{wb_metrics.ad_load_pct:.1f}%</code> — "
            "еженедельно отключайте кампании с ДРР выше целевого."
        )
    elif wb_metrics and 45 <= wb_metrics.buyout_coef_pct < 65:
        yellow_parts.append(
            f"выкуп <code>{wb_metrics.buyout_coef_pct:.1f}%</code> — "
            "подтяните инфографику, отзывы и размерную сетку."
        )
    loc_line = (prompt_metrics.localization_index_line or "").lower()
    if any(
        token in loc_line
        for token in ("низк", "ниже", "плох", "критич", "0.", "1.", "2.", "3.")
    ) and "не указан" not in loc_line:
        yellow_parts.append(
            "низкий индекс локализации — распределите остатки на региональные склады WB."
        )
    if not yellow_parts:
        yellow_parts.append("критических операционных отклонений не зафиксировано.")
    yellow += " ".join(yellow_parts)

    system_losses = float(wb_metrics.other_deductions if wb_metrics else 0.0)
    penalties_html = _build_penalties_traffic_text(
        system_losses,
        credit_deductions=credit,
        html=True,
    )
    red = penalties_html
    loss_display = agg.loss_sku_display_lines
    if loss_display or prompt_metrics.loss_sku_items:
        red_parts: list[str] = []
        if loss_display:
            red_parts.extend(line.removeprefix("• ") for line in loss_display)
        else:
            for item in prompt_metrics.loss_sku_items[:_TOP_LOSS_SKU_DISPLAY]:
                red_parts.append(
                    f"<b>{_escape_verdict(_oos_item_display_label(dict(item)))}</b> — убыток "
                    f"<code>{_fmt_rub_in_code(abs(float(item.get('net_profit_rub', 0.0))))}</code> руб."
                )
        if agg.tail_loss_sku_count > 0:
            red_parts.append(
                _tail_line_loss_skus(
                    agg.tail_loss_sku_count, agg.tail_loss_sku_rub
                ).removeprefix("• ")
            )
        if prompt_metrics.total_ad_cost > 0 and prompt_metrics.adv_load_pct >= 30:
            red_parts.append(
                f"Катастрофический ДРР <code>{prompt_metrics.adv_load_pct:.1f}%</code> — "
                f"реклама <code>{_fmt_rub_in_code(prompt_metrics.total_ad_cost)}</code> руб."
            )
        elif prompt_metrics.total_ad_cost > 0 and prompt_metrics.adv_load_pct > _HIGH_DRR_PCT:
            red_parts.append(
                f"ДРР <code>{prompt_metrics.adv_load_pct:.1f}%</code> — "
                "срочно режьте рекламный бюджет."
            )
        if red_parts:
            red = "🔴 <b>КРИТИЧЕСКАЯ ЗОНА:</b> " + "; ".join(red_parts)
    elif credit <= 0 and prompt_metrics.fomo_lost_rub <= 0 and not loss_display:
        red = f"🔴 <b>КРИТИЧЕСКАЯ ЗОНА:</b> {_build_penalties_traffic_text(0.0, html=False)}"

    return [green, "", yellow, "", red]


async def generate_wb_finance_consulting_html(
    settings: Settings,
    *,
    revenue_total: float,
    wb_metrics: WbMarketplaceMetrics | None = None,
    matrix_rows: list[list[str]] | None = None,
    models: list[str] | None = None,
    http_client: object | None = None,
    platform: str | None = None,
    file_path: str | Path | None = None,
    wb_json_str: str | None = None,
    tax_preset_id: str | None = None,
) -> str | None:
    """
    CFO-отчёт wb_ozon_finance: локальный HTML (мгновенно); OpenRouter — только по флагу.
    """
    import asyncio

    from content.chat_prompt import WB_ANALYTICS_SYSTEM_PROMPT

    revenue_total = float(revenue_total or 0.0)
    if revenue_total <= 0:
        revenue_total = await asyncio.to_thread(
            resolve_wb_revenue_total,
            calculated_total=0.0,
            file_path=file_path,
            matrix_rows=matrix_rows,
            platform=platform,
        )
    if revenue_total <= 0:
        return None

    if wb_metrics is None and matrix_rows:
        wb_metrics = resolve_wb_metrics_for_rows(
            matrix_rows, revenue_total, platform=platform
        )

    local_html = await asyncio.to_thread(
        _build_local_wb_finance_html,
        revenue_total,
        wb_metrics,
        matrix_rows=matrix_rows,
        platform=platform,
        tax_preset_id=tax_preset_id,
    )

    use_openrouter = bool(
        getattr(settings, "wb_finance_openrouter_html", False)
        and settings.openrouter_key
    )
    if not use_openrouter:
        return local_html

    ctx: dict[str, Any] | None = None
    if wb_json_str:
        try:
            parsed = json.loads(wb_json_str)
            if isinstance(parsed, dict) and not parsed.get("error"):
                ctx = parsed
        except json.JSONDecodeError:
            ctx = None
    if ctx is None:
        ctx = await asyncio.to_thread(
            resolve_wb_mpstats_context,
            file_path=file_path,
            matrix_rows=matrix_rows,
            revenue_total=revenue_total,
            platform=platform,
        )
    if not ctx or ctx.get("error"):
        return local_html

    model_chain: list[str] = []
    if models:
        model_chain.extend(m for m in models if m and m not in model_chain)
    for mid in settings.free_models:
        if mid and mid not in model_chain:
            model_chain.append(mid)

    messages = [
        {"role": "system", "content": WB_ANALYTICS_SYSTEM_PROMPT},
        {"role": "user", "content": build_wb_finance_json_user_message(ctx)},
    ]
    try:
        completion = await ask_ai_messages(
            settings,
            messages,
            timeout=min(12.0, settings.openrouter_timeout_sec),
            max_context_tokens=settings.chat_max_context_tokens_est,
            char_per_token=settings.chat_char_per_token_est,
            http_client=http_client,
            models=model_chain or None,
            max_tokens=settings.openrouter_premium_max_output_tokens,
            text_role="table_generator",
        )
        raw = sanitize_wb_finance_html(completion.get("content") or "")
        if _is_publishable_wb_finance_html(raw):
            if "cfo-v12" not in raw.lower():
                raw = f"{raw}\n\n<i>CFO build {_FINANCE_REPORT_BUILD}</i>"
            return append_wb_finance_mini_app_cta(normalize_finance_report_html(raw))
        logger.warning(
            "generate_wb_finance_consulting_html: OpenRouter HTML rejected, using local fallback"
        )
    except Exception:
        logger.exception("generate_wb_finance_consulting_html: OpenRouter failed")

    return local_html


def resolve_wb_metrics_for_rows(
    rows: list[list[str]],
    revenue_total: float,
    *,
    platform: str | None = None,
) -> WbMarketplaceMetrics | None:
    """Локальные метрики маркетплейса для промпта и fallback."""
    return compute_wb_marketplace_metrics(
        rows, revenue_total=revenue_total, platform=platform
    )
