"""Локальная сборка table_generator: HTML, Excel, график (JSON pipeline)."""

from __future__ import annotations

from io import BytesIO

import pytest

from services.table_generator_pack import (
    TABLE_XLSX_FILENAME,
    build_table_generator_pack,
    build_table_generator_pack_from_rows,
    build_xlsx_bytes,
    build_xlsx_bytes_from_table,
)
from services.table_mini_app_html import markdown_table_to_html_document
from services.table_json import parse_table_json_response

SAMPLE_JSON = (
    '{"title":"Доход","headers":["Месяц","Доход"],'
    '"rows":[["Янв",1200],["Фев",1500],["Мар",1800]]}'
)

SAMPLE_SHARE_JSON = (
    '{"title":"Доли","headers":["Категория","Доля"],'
    '"rows":[["Еда",40],["Транспорт",25],["Прочее",35]]}'
)


def test_parse_json_to_rows() -> None:
    payload = parse_table_json_response(SAMPLE_JSON)
    assert payload is not None
    rows = payload.to_rows_with_header()
    assert rows[0] == ["Месяц", "Доход"]
    assert rows[1] == ["Янв", "1200"]


def test_markdown_table_to_html_document() -> None:
    payload = parse_table_json_response(SAMPLE_JSON)
    assert payload is not None
    html = markdown_table_to_html_document(payload.to_rows_with_header(), title="Доход")
    assert "<!DOCTYPE html>" in html
    assert "tg-theme-bg-color" in html
    assert "position: sticky" in html
    assert "PAGE_SIZE" in html
    assert "Доход" in html
    assert "Месяц" in html
    assert "Янв" in html


def test_build_xlsx_bytes_from_table() -> None:
    payload = parse_table_json_response(SAMPLE_JSON)
    assert payload is not None
    data, total = build_xlsx_bytes_from_table(payload.headers, payload.rows)
    assert data[:2] == b"PK"
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data))
    ws = wb.active
    assert ws.cell(1, 1).value == "Месяц"
    assert ws.cell(2, 1).value == "Янв"


def test_build_xlsx_bytes_legacy_rows() -> None:
    rows = [["Месяц", "Доход"], ["Янв", "1200"]]
    data, total = build_xlsx_bytes(rows)
    assert total == 1200
    assert data[:2] == b"PK"


def test_build_xlsx_total_row_formula_and_styles() -> None:
    from openpyxl import load_workbook

    rows = [
        ["Месяц", "Выручка"],
        ["Январь", "60000"],
        ["Февраль", "55000"],
        ["Март", "70000"],
    ]
    data, total = build_xlsx_bytes(rows)
    assert total == 185_000
    wb = load_workbook(BytesIO(data))
    ws = wb.active

    assert ws.cell(1, 1).value == "Месяц"
    assert ws.cell(2, 2).value == 60000
    assert ws.cell(5, 1).value == "Итого"
    assert ws.cell(5, 2).value == 185_000
    assert ws.cell(5, 1).font.bold is True
    assert ws.cell(5, 2).font.bold is True
    assert ws.cell(5, 2).font.underline == "double"
    assert ws.cell(5, 1).fill.fill_type == "solid"


def test_xlsx_total_matches_telegram_one_screen() -> None:
    from services.table_json import parse_table_json_response
    from services.table_text_response import build_table_one_screen_html, compute_table_column_metrics

    raw = (
        '{"title":"Выручка","headers":["Месяц","Выручка"],'
        '"rows":[["Январь","60000"],["Февраль","55000"],["Март","70000"]]}'
    )
    payload = parse_table_json_response(raw)
    assert payload is not None
    metrics = compute_table_column_metrics(payload.to_rows_with_header())
    assert metrics is not None
    assert metrics.total == 185_000

    html = build_table_one_screen_html(payload, ai_insights="")
    assert "185,000" in html

    pack = build_table_generator_pack(raw, ai_insights="")
    assert pack is not None
    assert pack.calculated_total == 185_000
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(pack.xlsx_bytes))
    ws = wb.active
    assert ws.cell(5, 2).value == 185_000


def test_build_table_generator_pack_with_chart() -> None:
    pack = build_table_generator_pack(SAMPLE_JSON)
    assert pack is not None
    assert pack.chart_png_bytes is not None
    assert len(pack.chart_png_bytes) > 100
    assert pack.xlsx_bytes[:2] == b"PK"
    assert "Доход" in pack.telegram_caption_html
    assert TABLE_XLSX_FILENAME == "Отчет_Нейросеть.xlsx"


def test_build_table_generator_pack_no_numeric_chart() -> None:
    raw = '{"title":"X","headers":["A","B"],"rows":[["x","y"]]}'
    pack = build_table_generator_pack(raw)
    assert pack is not None
    assert pack.chart_png_bytes is None
    assert pack.xlsx_bytes


@pytest.mark.asyncio
async def test_run_chat_turn_table_returns_raw_json(repo_module, monkeypatch) -> None:
    from types import SimpleNamespace
    from unittest.mock import AsyncMock, patch

    from config import Settings
    from services.dialog_sanitize import compact_table_history_from_json
    from services.table_json import canonicalize_table_json
    from services.use_cases.chat_turn import ChatTurnOutcome, run_chat_turn
    from tests.conftest import TEST_ADMIN_IDS

    uid = TEST_ADMIN_IDS[0]
    await repo_module.ensure_user(uid)
    canonical = canonicalize_table_json(SAMPLE_JSON)
    assert canonical is not None

    fake_plan = SimpleNamespace(
        blocked=False,
        block_reason="",
        model_id="google/gemini-2.5-flash",
        max_tokens=1024,
        use_premium_prompt=True,
        energy_cost=20,
        crystal_cost=0,
    )
    billing_result = SimpleNamespace(
        plan=fake_plan,
        charge_id="chg",
        effective_role_id="table_generator",
        notice=None,
    )

    with patch(
        "services.use_cases.chat_turn.allow_request",
        new=AsyncMock(return_value=True),
    ), patch(
        "services.use_cases.chat_turn.billing.resolve_and_charge_text_chat",
        new=AsyncMock(return_value=billing_result),
    ), patch(
        "services.use_cases.chat_turn.dialog_append",
        new=AsyncMock(),
    ), patch(
        "services.use_cases.chat_turn.conv.build_openrouter_messages",
        new=AsyncMock(
            return_value=[{"role": "system", "content": "s"}, {"role": "user", "content": "u"}]
        ),
    ), patch(
        "services.use_cases.chat_turn.ask_ai_messages",
        new=AsyncMock(
            return_value={
                "content": SAMPLE_JSON,
                "prompt_tokens": 100,
                "completion_tokens": 50,
            }
        ),
    ) as ask_mock, patch(
        "services.use_cases.chat_turn.commit_assistant_turn_queued",
        new=AsyncMock(),
    ) as commit_mock, patch(
        "services.use_cases.chat_turn.insert_table_report",
        new=AsyncMock(return_value=42),
    ), patch(
        "services.use_cases.chat_turn.conv.schedule_memory_refresh",
    ):
        s = Settings(tg_token="x", openrouter_key="y", gemini_api_key="z")
        result = await run_chat_turn(s, uid, "сделай таблицу", text_role="table_generator")

    assert result.outcome is ChatTurnOutcome.SUCCESS
    assert result.table_raw_json == canonical
    assert result.assistant_message is None
    ask_mock.assert_awaited_once()
    assert ask_mock.await_args.kwargs.get("text_role") == "table_generator"
    commit_mock.assert_awaited_once()
    assert commit_mock.await_args.args[1] == compact_table_history_from_json(
        canonical,
        table_subrole="table_generator",
    )


def test_build_table_generator_pack_from_rows() -> None:
    rows = [["A", "B"], ["1", "2"]]
    pack = build_table_generator_pack_from_rows(rows, title="Тест")
    assert pack is not None
    assert pack.xlsx_bytes[:2] == b"PK"
