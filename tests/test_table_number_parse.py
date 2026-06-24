"""Безопасное приведение чисел в table_generator."""

from __future__ import annotations

from services.table_number_parse import coerce_excel_numeric, parse_table_number, prepare_excel_value, safe_float
from services.table_text_response import compute_table_column_metrics
from services.table_xlsx_flow import aggregate_wb_preview_products, build_wb_telegram_preview_html


def test_safe_float_from_int_and_float() -> None:
    assert safe_float(60000) == 60000.0
    assert safe_float(1200.5) == 1200.5


def test_safe_float_from_dirty_strings() -> None:
    assert safe_float("60 000 руб.") == 60000.0
    assert safe_float("60 000,50") == 60000.5
    assert safe_float("60,000.00") == 60000.0
    assert safe_float(None) == 0.0
    assert safe_float("abc") == 0.0


def test_wb_total_rub_from_string_cells() -> None:
    rows = [
        ["Наименование", "Выкупили, шт.", "Выручка"],
        ["Товар A", "2", "60 000 руб."],
        ["Товар B", "1", "70 000,50"],
    ]
    products = aggregate_wb_preview_products(rows)
    assert products is not None
    assert sum(p.rub for p in products) == 130_000.5

    html = build_wb_telegram_preview_html(rows, title="T")
    assert html is not None
    assert "0.00" not in html.split("ОБЩАЯ ВЫРУЧКА")[1][:40]


def test_metrics_from_numeric_json_cells() -> None:
    rows = [["Месяц", "Выручка"], ["Январь", 60000], ["Февраль", "55 000"]]
    metrics = compute_table_column_metrics(rows)
    assert metrics is not None
    assert metrics.total == 115_000.0


def test_prepare_excel_value_converts_dirty_strings() -> None:
    assert prepare_excel_value(60000) == 60000
    assert prepare_excel_value("60 000 руб.") == 60000
    assert prepare_excel_value("60 000,50") == 60000.5
    assert prepare_excel_value("Январь") == "Январь"
    assert prepare_excel_value(None) == ""


def test_excel_coerce_writes_numbers_not_strings() -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from services.table_generator_pack import build_xlsx_bytes

    rows = [["Месяц", "Выручка"], ["Январь", "60 000 руб."], ["Февраль", 55000]]
    data, total = build_xlsx_bytes(rows)
    assert total == 115_000
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    assert isinstance(ws.cell(2, 2).value, (int, float))
    assert ws.cell(2, 2).value == 60000
    assert isinstance(ws.cell(3, 2).value, (int, float))
    assert ws.cell(3, 2).value == 55000
    assert prepare_excel_value("60 000") == 60000
    assert coerce_excel_numeric("60 000") == 60000
    assert parse_table_number("  ") is None


def test_excel_precalculated_total_cell() -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from services.table_generator_pack import build_xlsx_bytes

    rows = [
        ["Месяц", "Выручка"],
        ["Январь", "60 000"],
        ["Февраль", "70 000"],
    ]
    data, total = build_xlsx_bytes(rows)
    assert total == 130_000
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    assert ws.cell(2, 2).value == 60000
    assert ws.cell(3, 2).value == 70000
    assert ws.cell(4, 2).value == 130_000
