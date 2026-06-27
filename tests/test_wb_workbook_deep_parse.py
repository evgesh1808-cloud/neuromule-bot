"""Глубокий многолистовой парсинг WB xlsx: Хранение, Удержания, маппинг складов."""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from services.file_processor import (
    collect_supply_chain_audit_from_rows,
    load_cfo_workbook_from_path,
    map_wb_warehouse_label,
    sync_table_cfo_processing_worker,
)


def _detail_headers() -> list[str]:
    return [
        "Предмет",
        "Артикул поставщика",
        "Склад",
        "Тип документа",
        "Обоснование для оплаты",
        "Кол-во",
        "Продажа (РРЦ)",
        "К перечислению продавцу за реализованный товар",
        "Услуги по доставке товара покупателю",
        "Вознаграждение Вайлдберриз",
    ]


def _write_multi_sheet_wb(path: Path, *, preamble: bool = False) -> None:
    wb = Workbook()
    ws_detail = wb.active
    ws_detail.title = "Детализация"
    if preamble:
        ws_detail.append(["Отчёт реализации Wildberries"])
        ws_detail.append(["Поставщик: ИП Тестов"])
        ws_detail.append([])
    ws_detail.append(_detail_headers())
    ws_detail.append(
        ["Товар", "SKU-1", "208547", "Продажа", "Продажа", "2", "2400", "1600", "100", "50"]
    )

    ws_storage = wb.create_sheet("Хранение")
    if preamble:
        ws_storage.append(["Платное хранение"])
        ws_storage.append([])
    ws_storage.append(["Склад", "Стоимость хранения, руб."])
    ws_storage.append(["208547", "2782.27"])

    ws_hold = wb.create_sheet("Удержания")
    if preamble:
        ws_hold.append(["Удержания WB"])
        ws_hold.append([])
    ws_hold.append(["Вид удержания", "Сумма удержания"])
    ws_hold.append(["Предоставление кредита", "8192.77"])

    wb.save(path)
    wb.close()


def test_map_wb_warehouse_label() -> None:
    assert map_wb_warehouse_label("208547") == "Рязань (Тюшевское)"
    assert map_wb_warehouse_label("50003969") == "Подольск (Транзит WB)"
    assert map_wb_warehouse_label("Казань") == "Казань"


def test_load_cfo_workbook_aux_sheets() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "wb_multi.xlsx"
        _write_multi_sheet_wb(path)
        loaded = load_cfo_workbook_from_path(str(path))

    assert len(loaded.matrix) >= 2
    assert loaded.aux_storage_cost == pytest.approx(2782.27)
    assert loaded.aux_system_losses == pytest.approx(8192.77)


def test_sync_worker_picks_aux_sheet_costs() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "wb_multi.xlsx"
        _write_multi_sheet_wb(path)
        result = sync_table_cfo_processing_worker(
            str(path),
            "wildberries",
            "wb_ozon_finance",
            "USN",
            6.0,
        )

    assert result["total_storage_cost"] == pytest.approx(2782.27)
    assert result["total_system_losses"] == pytest.approx(8192.77)
    assert result["top_warehouses"] == ["Рязань (Тюшевское)"]


def test_prompt_metrics_include_aux_sheet_costs_from_file() -> None:
    from services.file_processor import sync_table_cfo_processing_worker as cfo_etl_from_path
    from services.table_processing_worker import sync_table_processing_worker
    from services.table_wb_finance_ai import (
        build_wb_finance_express_html_local,
        compute_wb_finance_prompt_metrics,
    )

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "wb_multi.xlsx"
        _write_multi_sheet_wb(path)
        result = cfo_etl_from_path(
            str(path),
            "wildberries",
            "wb_ozon_finance",
            "USN",
            6.0,
        )
        worker = sync_table_processing_worker(
            str(path),
            "wb_ozon_finance",
            False,
            title="wb_multi",
            marketplace_platform="wildberries",
        )
        prompt = compute_wb_finance_prompt_metrics(
            float(result["total_revenue"]),
            None,
            matrix_rows=worker.rows if worker else None,
            aux_storage_cost=worker.aux_storage_cost if worker else 0.0,
            aux_system_losses=worker.aux_system_losses if worker else 0.0,
        )

    assert prompt is not None
    assert prompt.storage_cost == pytest.approx(2782.27)
    assert prompt.total_system_losses == pytest.approx(8192.77)

    html = build_wb_finance_express_html_local(prompt, None)
    assert "2,782.27" in html
    assert "8,192.77" in html


def test_load_cfo_workbook_aux_sheets_with_preamble() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "wb_preamble.xlsx"
        _write_multi_sheet_wb(path, preamble=True)
        loaded = load_cfo_workbook_from_path(str(path))

    assert loaded.aux_storage_cost == pytest.approx(2782.27)
    assert loaded.aux_system_losses == pytest.approx(8192.77)


def test_load_cfo_workbook_unnamed_aux_sheets() -> None:
    """Вкладки без слова «Хранение» в имени — по содержимому."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "wb_generic_tabs.xlsx"
        wb = Workbook()
        ws_detail = wb.active
        ws_detail.title = "Отчёт"
        ws_detail.append(_detail_headers())
        ws_detail.append(
            ["Товар", "SKU-1", "208547", "Продажа", "Продажа", "2", "2400", "1600", "100", "50"]
        )
        ws_storage = wb.create_sheet("Лист2")
        ws_storage.append(["Склад", "Стоимость хранения, руб."])
        ws_storage.append(["208547", "2782.27"])
        ws_hold = wb.create_sheet("Лист3")
        ws_hold.append(["Вид удержания", "Сумма удержания"])
        ws_hold.append(["Предоставление кредита", "8192.77"])
        wb.save(path)
        wb.close()
        loaded = load_cfo_workbook_from_path(str(path))

    assert loaded.aux_storage_cost == pytest.approx(2782.27)
    assert loaded.aux_system_losses == pytest.approx(8192.77)


def test_supply_chain_warehouse_id_mapping() -> None:
    matrix = [
        _detail_headers(),
        ["Товар", "SKU-1", "50003969", "Продажа", "Продажа", "1", "1200", "800", "50", "30"],
        ["Товар", "SKU-2", "208547", "Продажа", "Продажа", "1", "900", "600", "40", "20"],
    ]
    audit = collect_supply_chain_audit_from_rows(matrix)
    assert "Подольск (Транзит WB)" in audit["top_warehouses"]
    assert "Рязань (Тюшевское)" in audit["top_warehouses"]
